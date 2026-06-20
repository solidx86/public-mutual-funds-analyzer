import openpyxl, pytest


def _row(ws, r, name, abbr, shariah, ftype, rl, status, walpha, **kw):
    ws.cell(r, 1, name); ws.cell(r, 2, abbr); ws.cell(r, 3, shariah)
    ws.cell(r, 4, ftype); ws.cell(r, 6, rl); ws.cell(r, 10, status); ws.cell(r, 14, walpha)
    for col, val in kw.items():
        ws.cell(r, int(col[1:]), val)         # pass c35=.. style overrides


@pytest.fixture
def tiny_fundmaster(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Master"
    ws.cell(3, 1, "Fund Name")                 # header marker on row 3
    _row(ws, 4, "Public Index Fund", "PIX", "No", "Equity", 3, "Qualified", 2.1,
         c72=-3.0, c73=20)
    _row(ws, 5, "PB Growth Fund", "PBGF", "No", "Equity", 4, "Qualified", 1.0)   # PB → excluded
    _row(ws, 6, "Public e-Cash Deposit", "PeCDF-B", "No", "Money Market", 1, "Qualified", 0.1)  # -B → excluded
    _row(ws, 7, "Public Wholesale", "PWSIF", "No", "Equity", 5, "Qualified", 3.0)  # wholesale → excluded
    _row(ws, 8, "Public e-Cash Deposit", "PeCDF-A", "No", "Money Market", 1, "Qualified", 0.1)
    p = tmp_path / "PublicMutual_FundMaster_Jun2026_v0.1.0.xlsx"; wb.save(p)
    return str(p)
