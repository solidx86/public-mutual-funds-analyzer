"""Smoke tests for the screening pipeline, run fully offline from the tracked
cached JSONs (mfr_results.json, ath_results.json) — no PDFs or network needed."""

import csv
import json
import re

import openpyxl
import pytest

from conftest import REPO_ROOT, make_workspace, run_script

WA_WEIGHTS = {"ytd": 0.05, "1-year": 0.15, "3-year": 0.40, "5-year": 0.25, "10-year": 0.15}


def read_master_csv(workspace):
    with open(workspace / "data" / "cache" / "master_funds.csv", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ── End-to-end from tracked data ──────────────────────────────────────────────

def test_master_csv_covers_all_funds(pipeline_workspace):
    rows = read_master_csv(pipeline_workspace)
    source = json.loads((REPO_ROOT / "data" / "cache" / "mfr_results.json").read_text())
    assert len(rows) == len(source["all_funds"]) == 171

    statuses = {r["Status"] for r in rows}
    assert statuses == {"Qualified", "Disqualified"}
    qualified = sum(1 for r in rows if r["Status"] == "Qualified")
    assert qualified == 134

    for col in ("Fund Name", "Abbr", "Weighted Alpha (%)", "Risk Level", "Fund Type"):
        assert col in rows[0], f"missing column {col!r}"
    # risk levels joined from funds_risk_level.xlsx for nearly every fund
    assert sum(1 for r in rows if r["Risk Level"]) > 150


def test_qualification_matches_weighted_alpha(pipeline_workspace):
    """The qualifier is weighted alpha > 0 with >= 2 assessed periods — for every fund."""
    rows = read_master_csv(pipeline_workspace)
    for r in rows:
        wa = r["Weighted Alpha (%)"]
        total_p = int(r["Periods Assessed"].split("/")[1])
        expect = "Qualified" if wa and float(wa) > 0 and total_p >= 2 else "Disqualified"
        assert r["Status"] == expect, f"{r['Abbr']}: WA={wa!r} periods={total_p} -> {r['Status']}"


def test_workbook_shape_and_metadata(pipeline_workspace):
    skill_md = (REPO_ROOT / "fund-screener-skill" / "SKILL.md").read_text()
    version = re.search(r'^version:\s*["\']?([^"\'\n]+)', skill_md, re.M).group(1).strip()

    out = list((pipeline_workspace / "output" / "fundmasters").glob("*.xlsx"))
    assert len(out) == 1
    name = out[0].name
    assert re.fullmatch(
        rf"PublicMutual_FundMaster_[A-Z][a-z]{{2}}\d{{4}}_v{re.escape(version)}\.xlsx", name
    ), name

    wb = openpyxl.load_workbook(out[0])
    assert wb.sheetnames == ["Master", "Summary"]
    ws = wb["Master"]
    assert ws.max_column == 73
    fund_rows = [r for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value]
    assert len(fund_rows) == 171
    assert "134 Qualified / 171 Total Funds" in ws.cell(1, 1).value


# ── Weighted-alpha math on a hand-built fixture ───────────────────────────────

def synthetic_fund(abbr, alphas):
    """alphas: {period_key: alpha} — only listed periods have data."""
    detail = {k: {"fund": 1.0, "benchmark": 1.0 - v, "alpha": v} for k, v in alphas.items()}
    return {
        "name": f"SYNTHETIC {abbr} FUND",
        "abbr": abbr,
        "source_file": "[MFR APR26] Synthetic Series Funds.pdf",
        "asset_class": "Equity - Malaysia",
        "geography": "Malaysia",
        "is_shariah": False,
        "performance": {},
        "ytd": {},
        "period_detail": detail,
        "outperform_count": sum(1 for v in alphas.values() if v > 0),
        "total_periods": len(alphas),
        "outperform_rate": 0.0,
    }


def expected_wa(alphas):
    total_w = sum(WA_WEIGHTS[k] for k in alphas)
    return round(sum(WA_WEIGHTS[k] / total_w * v for k, v in alphas.items()), 4)


@pytest.fixture(scope="module")
def synthetic_rows(tmp_path_factory):
    full = {"ytd": 2.0, "1-year": 4.0, "3-year": 6.0, "5-year": -1.0, "10-year": 3.0}
    partial = {"1-year": 3.0, "3-year": -1.0}  # weights must renormalise to 0.15/0.55, 0.40/0.55
    negative = {"1-year": -2.0, "3-year": -4.0, "5-year": 1.0}
    sparse = {"3-year": 9.0}  # one period only -> no score, disqualified

    data = {
        "all_funds": [
            synthetic_fund("SYNFULL", full),
            synthetic_fund("SYNPART", partial),
            synthetic_fund("SYNNEG", negative),
            synthetic_fund("SYNONE", sparse),
        ],
        "qualified": [],
    }
    ws = make_workspace(tmp_path_factory.mktemp("synthetic"), data, with_ath=False)
    run_script(ws, "build_sheet_data.py")
    return {r["Abbr"]: r for r in read_master_csv(ws)}, {
        "SYNFULL": full, "SYNPART": partial, "SYNNEG": negative,
    }


def test_weighted_alpha_full_periods(synthetic_rows):
    rows, alphas = synthetic_rows
    # CSV renders floats to 2 decimals
    assert float(rows["SYNFULL"]["Weighted Alpha (%)"]) == round(expected_wa(alphas["SYNFULL"]), 2)
    assert rows["SYNFULL"]["Status"] == "Qualified"


def test_weighted_alpha_proportional_redistribution(synthetic_rows):
    rows, alphas = synthetic_rows
    # (0.15*3 + 0.40*-1) / 0.55 — missing periods redistribute proportionally
    assert expected_wa(alphas["SYNPART"]) == 0.0909
    assert float(rows["SYNPART"]["Weighted Alpha (%)"]) == 0.09
    assert rows["SYNPART"]["Status"] == "Qualified"


def test_negative_weighted_alpha_disqualifies(synthetic_rows):
    rows, alphas = synthetic_rows
    assert float(rows["SYNNEG"]["Weighted Alpha (%)"]) == round(expected_wa(alphas["SYNNEG"]), 2)
    assert float(rows["SYNNEG"]["Weighted Alpha (%)"]) < 0
    assert rows["SYNNEG"]["Status"] == "Disqualified"


def test_single_period_fund_has_no_score(synthetic_rows):
    rows, _ = synthetic_rows
    assert rows["SYNONE"]["Weighted Alpha (%)"] == ""
    assert rows["SYNONE"]["Status"] == "Disqualified"
