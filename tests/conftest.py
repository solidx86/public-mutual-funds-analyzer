"""Shared fixtures: build an isolated pipeline workspace under tmp so test runs
never touch the repo's tracked outputs. Scripts derive all paths from their own
location, so copying the skill bundle into the workspace redirects every read
and write there."""

import json
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "fund-screener-skill" / "scripts"


def make_workspace(dest: Path, mfr_data: dict, with_ath: bool = True) -> Path:
    """Lay out a minimal copy of the repo that the pipeline scripts can run in."""
    skill_dst = dest / "fund-screener-skill"
    skill_dst.mkdir(parents=True)
    shutil.copytree(SCRIPTS, skill_dst / "scripts")
    shutil.copy(REPO_ROOT / "fund-screener-skill" / "SKILL.md", skill_dst / "SKILL.md")

    # Pipeline scripts read/write cache files under data/cache and look up
    # funds_risk_level.xlsx under data/reference — mirror that layout here.
    cache_dir = dest / "data" / "cache"
    ref_dir = dest / "data" / "reference"
    cache_dir.mkdir(parents=True)
    ref_dir.mkdir(parents=True)

    ut_dir = dest / "unit-trust"
    ut_dir.mkdir()
    mfr_data = dict(mfr_data)
    mfr_data["base_dir"] = str(ut_dir)
    (cache_dir / "mfr_results.json").write_text(json.dumps(mfr_data))

    shutil.copy(REPO_ROOT / "data" / "reference" / "funds_risk_level.xlsx", ref_dir / "funds_risk_level.xlsx")
    if with_ath:
        shutil.copy(REPO_ROOT / "data" / "cache" / "ath_results.json", cache_dir / "ath_results.json")

    (dest / "output" / "fundmasters").mkdir(parents=True)
    return dest


def run_script(workspace: Path, script_name: str) -> subprocess.CompletedProcess:
    proc = subprocess.run(
        [sys.executable, str(workspace / "fund-screener-skill" / "scripts" / script_name)],
        cwd=workspace,
        capture_output=True,
        text=True,
        timeout=300,
    )
    assert proc.returncode == 0, f"{script_name} failed:\n{proc.stdout}\n{proc.stderr}"
    return proc


@pytest.fixture(scope="session")
def pipeline_workspace(tmp_path_factory):
    """Run the offline half of the pipeline (steps 3 and 4) once, from the
    tracked cached JSONs, exactly as a fresh clone would."""
    mfr_data = json.loads((REPO_ROOT / "data" / "cache" / "mfr_results.json").read_text())
    ws = make_workspace(tmp_path_factory.mktemp("pipeline"), mfr_data)
    run_script(ws, "build_sheet_data.py")
    run_script(ws, "build_xlsx.py")
    return ws


# ── consultant_engine (Track 0) fixtures ───────────────────────────────────
# These live in the root conftest rather than a tests/consultant_engine/conftest.py
# on purpose: that directory name collides with the production `consultant_engine`
# package (breaks pytest package import), and a second conftest.py would shadow the
# bare `from conftest import ...` that test_pipeline.py relies on. Root conftest
# fixtures are available to every test, including the consultant_engine suite.

def _row(ws, r, name, abbr, shariah, ftype, rl, status, walpha, **kw):
    ws.cell(r, 1, name); ws.cell(r, 2, abbr); ws.cell(r, 3, shariah)
    ws.cell(r, 4, ftype); ws.cell(r, 6, rl); ws.cell(r, 10, status); ws.cell(r, 14, walpha)
    for col, val in kw.items():
        ws.cell(r, int(col[1:]), val)         # pass c35=.. style overrides


@pytest.fixture
def tiny_fundmaster(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Master"
    ws.cell(3, 1, "Fund Name")                 # header marker on row 3
    _row(ws, 4, "Public Index Fund", "PIX", "No", "Equity", 3, "Qualified", 2.1,
         c72=-3.0, c73=20)
    _row(ws, 5, "PB Growth Fund", "PBGF", "No", "Equity", 4, "Qualified", 1.0)   # PB → excluded
    _row(ws, 6, "Public e-Cash Deposit", "PeCDF-B", "No", "Money Market", 1, "Qualified", 0.1)  # -B → excluded
    _row(ws, 7, "Public Wholesale", "PWSIF", "No", "Equity", 5, "Qualified", 3.0)  # wholesale → excluded
    _row(ws, 8, "Public e-Cash Deposit", "PeCDF-A", "No", "Money Market", 1, "Qualified", 0.1)
    p = tmp_path / "PublicMutual_FundMaster_Jun2026_v0.1.0.xlsx"; wb.save(p)
    return str(p)


@pytest.fixture
def fundmaster_4fund(tmp_path):
    """4-fund fixture for build_portfolio integration tests.

    Sheet "Master" (header row 3, data row 4+):
      - THREE conventional equity funds (RL 3, Qualified, Shariah No)
        each Equity-equivalent (dom_equity=75 for_equity=10) with distinct
        positive returns/alpha so they rank differently through CFS scoring.
      - PeEMAS (gold structural, RL 3, assets→other=95 for Balanced class).
      - PeCDF-A (money-market structural, RL 1, mm=100 for Defensive class).

    Equity funds use plain abbreviations (no "PB " prefix, no "-B" suffix,
    not in the WHOLESALE set) so _excluded() passes them all.
    """
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Master"
    ws.cell(3, 1, "Fund Name")   # header marker on row 3

    # ytd cols 15-17, 1y cols 18-20, 3y cols 21-23, 5y cols 24-26, 10y cols 27-29
    # col 72=drawdown, col 73=days_from_ath

    # Equity fund 1 — highest alpha
    _row(ws, 4, "Public Growth A", "PGA", "No", "Equity", 3, "Qualified", 4.0,
         c15=20.0, c16=15.0, c17=5.0,    # ytd: fund/bench/alpha
         c18=18.0, c19=13.0, c20=5.0,    # 1y
         c21=15.0, c22=11.0, c23=4.0,    # 3y: fund=15, bench=11, alpha=4
         c24=12.0, c25=9.0,  c26=3.0,    # 5y
         c27=10.0, c28=8.0,  c29=2.0,    # 10y
         c35=75.0, c36=10.0,             # dom_equity=75, for_equity=10 → Equity-equivalent
         c72=-5.0, c73=20)               # drawdown / days_from_ath for momentum

    # Equity fund 2 — middle alpha
    _row(ws, 5, "Public Balanced A", "PBA", "No", "Equity", 3, "Qualified", 3.0,
         c15=18.0, c16=14.0, c17=4.0,
         c18=16.0, c19=12.0, c20=4.0,
         c21=12.0, c22=9.0,  c23=3.0,
         c24=10.0, c25=8.0,  c26=2.5,
         c27=9.0,  c28=7.5,  c29=1.5,
         c35=75.0, c36=10.0,
         c72=-5.0, c73=30)

    # Equity fund 3 — lowest alpha of the three
    _row(ws, 6, "Public SmallCap A", "PSCA", "No", "Equity", 3, "Qualified", 2.0,
         c15=15.0, c16=12.0, c17=3.0,
         c18=13.0, c19=10.0, c20=3.0,
         c21=10.0, c22=8.0,  c23=2.0,
         c24=8.0,  c25=6.5,  c26=1.5,
         c27=7.0,  c28=6.0,  c29=1.0,
         c35=75.0, c36=10.0,
         c72=-5.0, c73=45)

    # PeEMAS — gold structural (RL 3; Moderate ceiling = 3 so no rl_ceiling issue;
    # other=95 → Balanced class so it won't compete with Equity-equivalent cores)
    _row(ws, 7, "Public e-Islamic EMAS", "PeEMAS", "No", "Gold", 3, "Qualified", 0.5,
         c15=5.0, c16=4.0, c17=1.0,
         c18=4.0, c19=3.5, c20=0.5,
         c21=3.0, c22=2.5, c23=0.5,
         c40=95.0,          # other=95 → Balanced (not Equity-equivalent)
         c72=-2.0, c73=10)

    # PeCDF-A — money-market structural (RL 1)
    _row(ws, 8, "Public e-Cash Deposit", "PeCDF-A", "No", "Money Market", 1, "Qualified", 0.1,
         c15=2.0, c16=2.0, c17=0.0,
         c18=2.0, c19=2.0, c20=0.0,
         c21=2.0, c22=2.0, c23=0.0,
         c38=100.0,          # mm=100 → Defensive
         c72=-1.0, c73=5)

    p = tmp_path / "PublicMutual_FundMaster_Jun2026_v0.1.0.xlsx"
    wb.save(p)
    return str(p)
