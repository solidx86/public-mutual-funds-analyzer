"""Regression checks for the LLM-generated proposals.

The fund-consultant skill is prompt-driven, so its output can drift in ways
unit tests on code never see. These tests pin the contract instead: every
proposal must conform to the locked template skeleton, its scoring arithmetic
must be internally consistent, and the disclosure rules must hold against the
FundMaster workbook the proposal says it was built from.

Scope: the tracked curated samples under output/examples/fund_proposals/ AND any
live run sitting in the gitignored output/fund_proposals/. In CI the live dir is
empty (fresh checkout), so this validates exactly the committed samples; locally
it doubles as a pre-flight on real runs — a generated proposal that trips a rule
fails the suite until it's fixed or moved out. Cited FundMaster workbooks are
resolved from output/examples/fundmasters/ or the live output/fundmasters/."""

import html
import re
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent

# Validate curated examples (tracked) and any local live run. Examples are
# listed first so a promoted-but-identical copy resolves to the live file when
# the same filename exists in both dirs.
_PROPOSAL_DIRS = (
    REPO_ROOT / "output" / "examples" / "fund_proposals",
    REPO_ROOT / "output" / "fund_proposals",
)
_FUNDMASTER_DIRS = (
    REPO_ROOT / "output" / "examples" / "fundmasters",
    REPO_ROOT / "output" / "fundmasters",
)


def _collect_proposals():
    by_name = {}
    for base in _PROPOSAL_DIRS:
        for p in base.glob("FundProposal_*.html"):
            by_name[p.name] = p  # later dir (live) wins on name collision
    return sorted(by_name.values(), key=lambda p: p.name)


PROPOSALS = _collect_proposals()

LOCKED_SECTIONS = [
    "Executive Summary",
    "Global &amp; Local Macro Context",
    "Client Risk Profile",
    "Fund Recommendations",
    "Portfolio Summary",
    "Portfolio Exposure",
    "Investment Strategy",
    "Fee Disclosure",
    "Disclaimer, Sources &amp; References",
]

WHOLESALE_ABBRS = {"PBCPF", "PWSIF", "PIWSIF", "PeWS20F"}

# Pinned violations found in earlier published samples (a Step-1b "PB "-fund
# recommendation, a renamed section, three CFS composites that didn't recompute).
# Those samples were regenerated clean in June 2026 — the sets stay so any NEW
# drift fails loudly.
KNOWN_ELIGIBILITY_VIOLATIONS = set()
KNOWN_SECTION_DEVIATIONS = {}
KNOWN_CFS_INCONSISTENCIES = set()


def fund_cards(text):
    """Split the Fund Recommendations section into per-fund card chunks."""
    chunks = re.split(r'<div class="fund-card">', text)[1:]
    cards = []
    for chunk in chunks:
        m = re.search(r"<h3>([^<]+)</h3>", chunk)
        title = html.unescape(m.group(1))
        # "Name · ABBR" with an optional " — role" suffix after the abbreviation
        abbr = title.rsplit("·", 1)[1].split("—")[0].strip()
        cards.append((abbr, chunk))
    return cards


def workbook_index(proposal_text):
    """Load the FundMaster workbook named on the proposal cover."""
    m = re.search(
        r'Data Source</div>\s*<div class="cover-meta-value">FundMaster (\w{3})\w* (\d{4})',
        proposal_text,
    )
    assert m, "cover must name its FundMaster data source"
    pattern = f"PublicMutual_FundMaster_{m.group(1)}{m.group(2)}_v*.xlsx"
    matches = [p for d in _FUNDMASTER_DIRS for p in d.glob(pattern)]
    assert matches, (
        f"workbook for {m.group(1)} {m.group(2)} not found in "
        "output/examples/fundmasters/ or output/fundmasters/"
    )
    ws = openpyxl.load_workbook(matches[0])["Master"]
    funds = {}
    for r in range(4, ws.max_row + 1):
        abbr = ws.cell(r, 2).value
        if abbr:
            funds[abbr] = {"name": ws.cell(r, 1).value, "status": ws.cell(r, 10).value}
    return funds


def skill_version():
    text = (REPO_ROOT / "fund-consultant-skill" / "SKILL.md").read_text()
    return re.search(r'^version:\s*["\']?([^"\'\n]+)', text, re.M).group(1).strip()


def test_proposal_samples_exist():
    assert PROPOSALS, "no tracked sample proposals found"


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_locked_template_sections_in_order(path):
    text = path.read_text()
    titles = re.findall(r'<div class="section-title">([^<]+)</div>', text)
    expected = list(LOCKED_SECTIONS)
    for idx, title in KNOWN_SECTION_DEVIATIONS.get(path.name, {}).items():
        expected[idx] = title
    assert titles == expected, f"section skeleton drifted: {titles}"
    assert '<div class="cover">' in text


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_version_stamp_and_ai_disclosure(path):
    text = path.read_text()
    ver = skill_version()
    assert path.name.endswith(f"_v{ver}.html"), "filename must carry the skill version"
    assert f"fund-consultant v{ver}" in text, "cover footer must carry the skill version"
    assert "AI-Generated Document" in text, "AI-generation disclaimer block missing"


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_cfs_scores_are_internally_consistent(path):
    """Each CFS bar: composite in (0,100], four dimensions in [0,100], weights sum to 100."""
    text = path.read_text()
    bars = re.split(r'<div class="cfs-bar">', text)[1:]
    assert bars, "no CFS bars found in proposal"
    for bar in bars:
        bar = bar.split('<div class="cfs-bar-note"')[0]
        score_m = re.search(r'class="cfs-score">([\d.]+)</span>', bar)
        if score_m is None:
            # structural/passive holdings (e.g. gold tracker) show "n/a" or "—"
            title = html.unescape(bar.split("</div>")[0])
            assert "n/a" in title or "—" in title, f"CFS bar without score: {bar[:120]!r}"
            continue
        composite = float(score_m.group(1))
        assert 0 < composite <= 100
        # "<score> / 100 · <w>% weight" — score may carry a footnote asterisk
        rows = re.findall(r"(\d+(?:\.\d+)?) / 100\*? &middot; (\d+(?:\.\d+)?)% weight", bar)
        assert len(rows) == 4, "CFS bar must show all four dimensions"
        scores = [float(s) for s, _ in rows]
        weights = [float(w) for _, w in rows]
        assert all(0 <= s <= 100 for s in scores)
        assert sum(weights) == pytest.approx(100, abs=0.5)
        # composite must equal the weighted sum of its displayed dimensions
        recomputed = sum(s * w / 100 for s, w in zip(scores, weights))
        if (path.name, composite) in KNOWN_CFS_INCONSISTENCIES:
            continue
        assert composite == pytest.approx(recomputed, abs=1.0), (
            f"CFS {composite} != recomputed {recomputed:.1f}"
        )


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_recommended_funds_exist_in_source_workbook(path):
    text = path.read_text()
    funds = workbook_index(text)
    cards = fund_cards(text)
    assert cards, "no fund cards found"
    for abbr, _ in cards:
        assert abbr in funds, f"recommended fund {abbr} not in source FundMaster"


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_alpha_warning_iff_disqualified(path):
    """Disclosure rule: a fund card carries an ALPHA WARNING block exactly when
    the fund failed screening (weighted alpha <= 0) in the source workbook."""
    text = path.read_text()
    funds = workbook_index(text)
    for abbr, chunk in fund_cards(text):
        has_warning = '<div class="alpha-warning">' in chunk
        disqualified = funds[abbr]["status"] == "Disqualified"
        assert has_warning == disqualified, (
            f"{abbr}: status={funds[abbr]['status']}, alpha-warning present={has_warning}"
        )


@pytest.mark.parametrize("path", PROPOSALS, ids=lambda p: p.name)
def test_no_new_retail_eligibility_violations(path):
    """SKILL.md Step 1b: no 'PB '-named funds, no '-B' class units, no wholesale
    funds may be recommended. Existing violations are pinned above."""
    text = path.read_text()
    funds = workbook_index(text)
    violations = set()
    for abbr, _ in fund_cards(text):
        name = funds[abbr]["name"] or ""
        if name.startswith("PB ") or abbr.endswith("-B") or abbr in WHOLESALE_ABBRS:
            violations.add((path.name, abbr))
    new = violations - KNOWN_ELIGIBILITY_VIOLATIONS
    assert not new, f"new retail-eligibility violations: {new}"
