"""Shared validation rules for generated fund proposals.

Pure functions that return ``list[dict]`` of ``{"code": str, "msg": str}``.
Empty list means clean — no violations.

Both consumers import from here:
  * runtime ``validate`` node in the consultant engine graph
  * offline pytest suite (tests/consultant_engine/test_validation_rules.py and
    the legacy tests/test_proposal_validation.py, unchanged)
"""

from __future__ import annotations

import html as _html
import re
from pathlib import Path
from typing import Any

import openpyxl

# ── Shared constants ─────────────────────────────────────────────────────────

LOCKED_SECTIONS: list[str] = [
    "Executive Summary",
    "Global &amp; Local Macro Context",
    "Client Risk Profile",
    "Fund Recommendations",
    "Portfolio Summary",
    "Portfolio Exposure",
    "Investment Strategy",
    "Fee Disclosure",
    "Disclaimer, Sources &amp; References",
]

WHOLESALE_ABBRS: frozenset[str] = frozenset({"PBCPF", "PWSIF", "PIWSIF", "PeWS20F"})

_DISCLOSURE_H4S: list[str] = [
    "AI-Generated Document",
    "Regulatory Disclaimer",
    "Cooling-Off Right",
    "Conflict of Interest",
]

# ── Shared helpers ────────────────────────────────────────────────────────────


def fund_cards(text: str) -> list[tuple[str, str]]:
    """Split the Fund Recommendations section into per-fund card chunks.

    Returns a list of ``(abbr, chunk_html)`` tuples.
    """
    chunks = re.split(r'<div class="fund-card">', text)[1:]
    cards: list[tuple[str, str]] = []
    for chunk in chunks:
        m = re.search(r"<h3>([^<]+)</h3>", chunk)
        if not m:
            continue
        title = _html.unescape(m.group(1))
        # "Name · ABBR" with an optional " — role" suffix after the abbreviation
        abbr = title.rsplit("·", 1)[1].split("—")[0].strip()
        cards.append((abbr, chunk))
    return cards


def workbook_index(workbook_path: str | Path) -> dict[str, dict[str, Any]]:
    """Load a FundMaster workbook and return an abbr-keyed index.

    Each entry: ``{"name": str, "status": str}``.
    The ``status`` column (col 10) holds "Qualified" | "Disqualified".
    """
    ws = openpyxl.load_workbook(str(workbook_path))["Master"]
    funds: dict[str, dict[str, Any]] = {}
    for r in range(4, ws.max_row + 1):
        abbr = ws.cell(r, 2).value
        if abbr:
            funds[abbr] = {
                "name": ws.cell(r, 1).value,
                "status": ws.cell(r, 10).value,
            }
    return funds


# ── Individual rule functions ─────────────────────────────────────────────────


def check_sections(html_text: str) -> list[dict[str, str]]:
    """Verify the 9 locked section titles are present, in order, and that
    there are exactly 9 ``.section`` divs.

    Violation codes:
      ``section_count``  — wrong number of .section divs
      ``section_order``  — section title out of order / missing / extra
    """
    violations: list[dict[str, str]] = []

    # Count <div class="section"> occurrences
    section_div_count = len(re.findall(r'<div class="section">', html_text))
    if section_div_count != 9:
        violations.append({
            "code": "section_count",
            "msg": (
                f"Expected 9 .section divs, found {section_div_count}"
            ),
        })

    # Extract ordered section titles
    titles = re.findall(r'<div class="section-title">([^<]+)</div>', html_text)
    if titles != LOCKED_SECTIONS:
        # Determine specific drift
        if set(titles) == set(LOCKED_SECTIONS) and titles != LOCKED_SECTIONS:
            violations.append({
                "code": "section_order",
                "msg": f"Sections present but out of order: {titles}",
            })
        else:
            # Could be missing, extra, or both — still report as section_order
            violations.append({
                "code": "section_order",
                "msg": f"Section skeleton drifted: {titles}",
            })

    return violations


def check_version_and_disclosure(html_text: str, version: str) -> list[dict[str, str]]:
    """Verify version stamp and the 4 mandatory disclosure headings.

    Violation codes:
      ``skill_version_literal``  — unresolved ``[SKILL_VERSION]`` placeholder
      ``version_mismatch``       — cover footer version doesn't match ``version``
      ``disclosure_heading``     — one or more of the 4 required h4s missing/out of order
    """
    violations: list[dict[str, str]] = []

    if "[SKILL_VERSION]" in html_text:
        violations.append({
            "code": "skill_version_literal",
            "msg": "Unresolved [SKILL_VERSION] literal found in proposal",
        })

    if f"fund-consultant v{version}" not in html_text:
        # Find what version is actually stamped
        m = re.search(r"fund-consultant v([\d.]+)", html_text)
        found = m.group(1) if m else "<none>"
        violations.append({
            "code": "version_mismatch",
            "msg": (
                f"Expected 'fund-consultant v{version}' in cover footer, "
                f"found 'fund-consultant v{found}'"
            ),
        })

    # Check all 4 disclosure headings present in order (within the disclaimer section)
    h4_headings = re.findall(r"<h4>([^<]+)</h4>", html_text)
    # Look for all 4 in order within h4_headings
    found_order: list[str] = []
    h4_iter = iter(h4_headings)
    for required in _DISCLOSURE_H4S:
        for heading in h4_iter:
            if heading.strip() == required:
                found_order.append(heading.strip())
                break
    if found_order != _DISCLOSURE_H4S:
        missing = [h for h in _DISCLOSURE_H4S if h not in found_order]
        violations.append({
            "code": "disclosure_heading",
            "msg": f"Disclosure headings missing or out of order: {missing}",
        })

    return violations


def check_cfs_consistency(html_text: str) -> list[dict[str, str]]:
    """Recompute each fund card's CFS composite from its displayed dimension
    scores + weights and compare to the displayed composite.

    Bounded per fund card via ``fund_cards`` so each recompute reads exactly one
    card's ``cfs-score`` and its four dimension rows — the same per-card iteration
    the summary / render-fidelity checks use, instead of an unbounded split on
    ``cfs-bar`` whose segments leaked across cards. Mismatch tolerance: ±1.0.

    A card with **no** ``cfs-score`` is a structural / passive holding (gold, money
    market) and is legitimately skipped. But a card that HAS a ``cfs-score`` yet
    does not render exactly four dimension rows breaks the scored-card contract and
    **fails loud** (``malformed_cfs_bar``) rather than being silently skipped — a
    silent skip there would let a malformed bar pass the validator unchecked.

    Violation codes: ``cfs_recompute``, ``malformed_cfs_bar``
    """
    violations: list[dict[str, str]] = []

    for abbr, chunk in fund_cards(html_text):
        score_m = re.search(r'class="cfs-score">([\d.]+)</span>', chunk)
        if score_m is None:
            continue  # structural / passive holding — no composite to reconcile
        composite = float(score_m.group(1))
        rows = re.findall(
            r"(\d+(?:\.\d+)?) / 100\*? &middot; (\d+(?:\.\d+)?)% weight",
            chunk,
        )
        if len(rows) != 4:
            violations.append({
                "code": "malformed_cfs_bar",
                "msg": (
                    f"{abbr}: scored card renders {len(rows)} CFS dimension rows, "
                    f"expected 4 — composite {composite} cannot be reconciled"
                ),
            })
            continue
        scores = [float(s) for s, _ in rows]
        weights = [float(w) for _, w in rows]
        recomputed = sum(s * w / 100 for s, w in zip(scores, weights))
        if abs(composite - recomputed) > 1.0:
            violations.append({
                "code": "cfs_recompute",
                "msg": (
                    f"{abbr}: CFS composite {composite} != recomputed "
                    f"{recomputed:.2f} (diff {abs(composite - recomputed):.2f})"
                ),
            })

    return violations


def check_perf_consistency(html_text: str) -> list[dict[str, str]]:
    """For each performance-table row, verify displayed Alpha == Fund - Bench
    (the MFR 'value add' relation). Rows with any missing cell ("&mdash;"/"—")
    are skipped. Tolerance: ±0.1.

    Violation code: ``perf_recompute``
    """
    violations: list[dict[str, str]] = []
    tables = re.findall(r'<table class="perf-table">.*?</table>', html_text, re.DOTALL)
    for table in tables:
        for row in re.findall(r"<tr>(.*?)</tr>", table, re.DOTALL):
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL)
            if len(cells) != 4:
                continue

            def _num(cell: str):
                txt = _html.unescape(re.sub(r"<[^>]+>", "", cell)).strip().replace("+", "")
                if txt in ("—", "-", "", "n/a", "N/A"):
                    return None
                try:
                    return float(txt)
                except ValueError:
                    return None

            period = _html.unescape(re.sub(r"<[^>]+>", "", cells[0])).strip()
            fund, bench, alpha = _num(cells[1]), _num(cells[2]), _num(cells[3])
            if None in (fund, bench, alpha):
                continue
            if abs(alpha - (fund - bench)) > 0.1:
                violations.append({
                    "code": "perf_recompute",
                    "msg": (
                        f"{period}: displayed alpha {alpha} != fund-bench "
                        f"{fund - bench:.2f} (diff {abs(alpha - (fund - bench)):.2f})"
                    ),
                })
    return violations


def check_exposure_consistency(html_text: str) -> list[dict[str, str]]:
    """Within the Portfolio Exposure section, verify each ``.exposure-chart-block``
    legend's ``legend-pct`` values sum to 100.0 (a pie is parts-of-a-whole).

    Scoped to the exposure section so it never picks up unrelated ``legend-pct``
    spans elsewhere in the document. Tolerance: ±2.0.

    Violation code: ``exposure_sum``
    """
    violations: list[dict[str, str]] = []

    # Isolate the Portfolio Exposure section (up to the next section or EOF).
    sect_m = re.search(
        r'Portfolio Exposure</div>(.*?)(?=<div class="section">|</body>)',
        html_text,
        re.DOTALL,
    )
    if sect_m is None:
        return violations  # section absent — check_sections owns that failure
    section = sect_m.group(1)

    blocks = re.findall(
        r'<div class="exposure-chart-block">(.*?)(?=<div class="exposure-chart-block">|$)',
        section,
        re.DOTALL,
    )
    for block in blocks:
        title_m = re.search(r'exposure-chart-title">([^<]+)<', block)
        title = title_m.group(1).strip() if title_m else "?"
        total = 0.0
        found = False
        for raw in re.findall(r'class="legend-pct"[^>]*>([^<]*)<', block):
            txt = _html.unescape(re.sub(r"<[^>]+>", "", raw)).strip()
            txt = txt.replace("%", "").replace("+", "")
            try:
                total += float(txt)
                found = True
            except ValueError:
                continue
        if not found:
            continue  # no numeric legend in this block — nothing to check
        if abs(total - 100.0) > 2.0:
            violations.append({
                "code": "exposure_sum",
                "msg": (
                    f"{title} exposure legend sums to {total:.1f} "
                    f"(expected 100.0 ± 2.0)"
                ),
            })
    return violations


def check_summary_consistency(html_text: str) -> list[dict[str, str]]:
    """Cross-check each Portfolio Summary row's CFS against the same fund's
    fund-card composite.

    Both numbers are Python-rendered from the same ``cfs_scores``, so a mismatch
    means the summary table was corrupted downstream (defense-in-depth for the
    Python-owned summary numbers, mirroring the CFS/perf/exposure guards).
    Structural rows (no card composite) and ``—`` cells are skipped.
    Tolerance: ±1.0.

    Violation code: ``summary_mismatch``
    """
    violations: list[dict[str, str]] = []

    # Composite per fund from the fund cards (the cfs-score span).
    card_cfs: dict[str, float] = {}
    for abbr, chunk in fund_cards(html_text):
        m = re.search(r'class="cfs-score">([\d.]+)</span>', chunk)
        if m:
            card_cfs[abbr] = float(m.group(1))

    # Isolate the Portfolio Summary section (up to the next section or EOF).
    sect_m = re.search(
        r'Portfolio Summary</div>(.*?)(?=<div class="section">|</body>)',
        html_text,
        re.DOTALL,
    )
    if sect_m is None:
        return violations

    for row in re.findall(r"<tr>(.*?)</tr>", sect_m.group(1), re.DOTALL):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL)
        if len(cells) < 4:
            continue
        abbr = _html.unescape(re.sub(r"<[^>]+>", "", cells[0])).strip()
        if abbr not in card_cfs:
            continue  # structural row / fund without a card composite
        cfs_txt = _html.unescape(re.sub(r"<[^>]+>", "", cells[3])).strip()
        try:
            summary_cfs = float(cfs_txt)
        except ValueError:
            continue  # "—" or non-numeric
        if abs(summary_cfs - card_cfs[abbr]) > 1.0:
            violations.append({
                "code": "summary_mismatch",
                "msg": (
                    f"{abbr}: Portfolio Summary CFS {summary_cfs} != "
                    f"fund-card composite {card_cfs[abbr]}"
                ),
            })
    return violations


def check_funds_in_workbook(
    html_text: str,
    wb_index: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """Verify every fund card abbreviation exists in the workbook index.

    Violation code: ``fund_not_in_workbook``
    """
    violations: list[dict[str, str]] = []
    for abbr, _ in fund_cards(html_text):
        if abbr not in wb_index:
            violations.append({
                "code": "fund_not_in_workbook",
                "msg": f"Recommended fund '{abbr}' not found in source FundMaster workbook",
            })
    return violations


def check_alpha_warning(
    html_text: str,
    wb_index: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """Verify alpha-warning blocks appear exactly when a fund is Disqualified.

    Disclosure rule: a fund card carries an ALPHA WARNING block iff
    the fund's status in the source workbook is "Disqualified".

    Violation code: ``alpha_warning``
    """
    violations: list[dict[str, str]] = []
    for abbr, chunk in fund_cards(html_text):
        if abbr not in wb_index:
            continue  # already caught by check_funds_in_workbook
        has_warning = '<div class="alpha-warning">' in chunk
        disqualified = wb_index[abbr]["status"] == "Disqualified"
        if has_warning != disqualified:
            violations.append({
                "code": "alpha_warning",
                "msg": (
                    f"{abbr}: status={wb_index[abbr]['status']}, "
                    f"alpha-warning present={has_warning}"
                ),
            })
    return violations


def check_retail_eligibility(
    html_text: str,
    wb_index: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """Verify no PB-named, -B class, or wholesale funds are recommended.

    Violation code: ``retail_eligibility``
    """
    violations: list[dict[str, str]] = []
    for abbr, _ in fund_cards(html_text):
        if abbr not in wb_index:
            continue  # already caught by check_funds_in_workbook
        name = wb_index[abbr]["name"] or ""
        if name.startswith("PB ") or abbr.endswith("-B") or abbr in WHOLESALE_ABBRS:
            violations.append({
                "code": "retail_eligibility",
                "msg": (
                    f"{abbr} ({name!r}) fails retail eligibility: "
                    "PB-named, -B class, or wholesale fund"
                ),
            })
    return violations


# ── Unfilled prose slot guard ─────────────────────────────────────────────────

# Two leak forms, both of which mean a prose slot never received real content.
# The fake-LLM convention ``[KEY narrative]`` is deliberately NOT listed: offline/
# eval runs fill every slot with that readable placeholder by design, so it must
# stay valid. A genuine real-LLM miss degrades to the distinct ``[UNFILLED:KEY]``
# sentinel instead, which is what we catch here.
_UNFILLED_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"<!--slot:([^>]+?)-->"),
     "raw <!--slot:{key}--> marker survived prose fill"),
    (re.compile(r"\[UNFILLED:([^\]]+)\]"),
     "prose slot '{key}' fell back to an unfilled sentinel"),
]


def check_unfilled_slots(html_text: str) -> list[dict[str, str]]:
    """Verify no prose slot leaked into the rendered proposal.

    Catches two failure forms:
      * a raw ``<!--slot:KEY-->`` marker that survived substitution (a structural
        fill bug), and
      * a ``[UNFILLED:KEY]`` sentinel emitted when the LLM prose fill — even after
        its one targeted retry — never returned that key.

    The fake-LLM convention ``[KEY narrative]`` is intentionally NOT matched: in
    offline/eval runs every slot is a readable placeholder by design.

    Violation code: ``unfilled_slot``
    """
    violations: list[dict[str, str]] = []
    seen: set[str] = set()
    for pattern, template in _UNFILLED_PATTERNS:
        for m in pattern.finditer(html_text):
            msg = template.format(key=m.group(1).strip())
            if msg in seen:
                continue
            seen.add(msg)
            violations.append({"code": "unfilled_slot", "msg": msg})
    return violations


# ── Render-fidelity reconciliation (state-aware; runtime validate node only) ──
#
# The repair pass round-trips the WHOLE document through the LLM ("return the
# complete HTML"), so the model CAN corrupt a Python-owned value (a CFS number, a
# Shariah label, an allocation %). This check reconciles the rendered Python-owned
# values in the FINAL html against the engine's OWN authoritative state.
#
# Reconciliation, NOT re-derivation: every authoritative value comes from either a
# value already stored in state (cfs_scores / portfolio / eligible_funds) OR by
# REUSING the exact production compute function the engine itself used
# (compute_asset_exposure / compute_geo_exposure / _compute_portfolio_metrics). No
# formula is re-implemented here — re-coding the math would be a circular test.

# Number renders at display precision; ±0.05 absorbs 1-dp rounding without masking
# a real corruption (e.g. 79.4 → 66.4).
_FIDELITY_TOL = 0.05

# Cells that legitimately render with no authoritative value to compare against.
_FIDELITY_BLANK = {"", "—", "&mdash;", "n/a", "N/A", "-"}


def _fidelity_num(raw: str):
    """Parse a rendered numeric cell to float, or None when blank/non-numeric.

    Strips tags, unescapes entities, and drops ``%`` / ``+`` / ``RM`` / commas so
    "+2.40%", "RM 1,000" and "79.4" all parse. Em-dash / empty → None (skip, never
    false-fire)."""
    txt = _html.unescape(re.sub(r"<[^>]+>", "", raw)).strip()
    if txt in _FIDELITY_BLANK:
        return None
    txt = txt.replace("%", "").replace("+", "").replace(",", "").replace("RM", "").strip()
    try:
        return float(txt)
    except ValueError:
        return None


def _cfs_by_abbr(state: Any) -> dict[str, dict]:
    return {c["abbr"]: c for c in state.get("cfs_scores", []) if c.get("abbr")}


def _fund_by_abbr(state: Any) -> dict[str, dict]:
    return {f["abbr"]: f for f in state.get("eligible_funds", []) if f.get("abbr")}


def _alloc_by_abbr(state: Any) -> dict[str, float]:
    return {h["abbr"]: h["allocation_pct"] for h in state.get("portfolio", []) if h.get("abbr")}


def _shariah_label(value) -> str:
    """Map a fund's stored ``shariah`` bool to its rendered card label."""
    return "Shariah" if value else "Conventional"


def _check_card_coverage(html_text: str, state: Any) -> list[dict[str, str]]:
    """Coverage canary: every portfolio holding in state must render as a fund card.

    The other fidelity checks iterate the cards they FIND — a card that was dropped
    or whose abbreviation was mangled simply isn't visited, so its values would pass
    unchecked (the classic vacuous pass). This asserts the rendered card set covers
    every holding in ``state["portfolio"]``, turning a missing/renamed card into a
    loud failure instead of a silent skip.
    """
    violations: list[dict[str, str]] = []
    expected = [h["abbr"] for h in state.get("portfolio", []) if h.get("abbr")]
    rendered = {abbr for abbr, _ in fund_cards(html_text)}
    for abbr in expected:
        if abbr not in rendered:
            violations.append({
                "code": "RENDER_FIDELITY",
                "msg": (
                    f"{abbr}: portfolio holding has no fund card in the rendered "
                    f"proposal (expected {len(expected)} holdings, found "
                    f"{len(rendered)} cards)"
                ),
            })
    return violations


def _check_fund_card_fidelity(html_text: str, state: Any) -> list[dict[str, str]]:
    """Reconcile each fund card's rendered CFS composite, allocation, and Shariah
    label against the authoritative state values for that fund."""
    violations: list[dict[str, str]] = []
    cfs = _cfs_by_abbr(state)
    funds = _fund_by_abbr(state)
    allocs = _alloc_by_abbr(state)

    for abbr, chunk in fund_cards(html_text):
        # CFS composite (cores only — structural cards carry no cfs-score span).
        score_m = re.search(r'class="cfs-score">([\d.]+)</span>', chunk)
        if score_m is not None and abbr in cfs:
            rendered = float(score_m.group(1))
            authoritative = cfs[abbr].get("composite")
            if authoritative is not None and abs(rendered - float(authoritative)) > _FIDELITY_TOL:
                violations.append({
                    "code": "RENDER_FIDELITY",
                    "msg": (
                        f"{abbr}: rendered CFS composite {rendered} != engine "
                        f"value {authoritative}"
                    ),
                })

        # Allocation % (fund-card header).
        alloc_m = re.search(r'class="alloc">([\d.]+)%</span>', chunk)
        if alloc_m is not None and abbr in allocs:
            rendered = float(alloc_m.group(1))
            if abs(rendered - float(allocs[abbr])) > _FIDELITY_TOL:
                violations.append({
                    "code": "RENDER_FIDELITY",
                    "msg": (
                        f"{abbr}: rendered allocation {rendered}% != engine value "
                        f"{allocs[abbr]}%"
                    ),
                })

        if abbr not in funds:
            continue
        fund = funds[abbr]

        # Shariah label (state stores a bool; HTML renders Shariah/Conventional).
        sh_m = re.search(r"<strong>Shariah:</strong>\s*([A-Za-z]+)", chunk)
        if sh_m is not None and "shariah" in fund:
            rendered = sh_m.group(1).strip()
            expected = _shariah_label(fund.get("shariah"))
            if rendered != expected:
                violations.append({
                    "code": "RENDER_FIDELITY",
                    "msg": f"{abbr}: rendered Shariah label '{rendered}' != engine '{expected}'",
                })

        # Risk Level meta cell.
        rl_m = re.search(r"<strong>RL:</strong>\s*([^<]+)", chunk)
        if rl_m is not None and fund.get("risk_level") is not None:
            rendered = _fidelity_num(rl_m.group(1))
            if rendered is not None and abs(rendered - float(fund["risk_level"])) > _FIDELITY_TOL:
                violations.append({
                    "code": "RENDER_FIDELITY",
                    "msg": (
                        f"{abbr}: rendered Risk Level {rendered} != engine "
                        f"{fund['risk_level']}"
                    ),
                })

        # Volatility Factor meta cell (skipped gracefully when rendered as em-dash).
        vf_m = re.search(r"<strong>VF:</strong>\s*([^<]+)", chunk)
        if vf_m is not None and fund.get("volatility_factor") is not None:
            rendered = _fidelity_num(vf_m.group(1))
            if rendered is not None and abs(rendered - float(fund["volatility_factor"])) > _FIDELITY_TOL:
                violations.append({
                    "code": "RENDER_FIDELITY",
                    "msg": (
                        f"{abbr}: rendered Volatility Factor {rendered} != engine "
                        f"{fund['volatility_factor']}"
                    ),
                })

    return violations


def _check_exposure_fidelity(html_text: str, state: Any) -> list[dict[str, str]]:
    """Reconcile each exposure legend % against the authoritative value from the
    production look-through functions (compute_asset_exposure / compute_geo_exposure).

    Matches by legend LABEL so a swapped pair (sum still 100, individual values
    wrong) is caught — something check_exposure_sum cannot see.
    """
    from consultant_engine.exposure import (  # local: avoid any import-cycle risk
        _ASSET_LABELS,
        compute_asset_exposure,
        compute_geo_exposure,
    )

    violations: list[dict[str, str]] = []
    portfolio = state.get("portfolio", [])
    funds_by_abbr = {f["abbr"]: f for f in state.get("eligible_funds", []) if f.get("abbr")}

    # Authoritative percentage per legend label, from the production functions.
    authoritative: dict[str, float] = {}
    asset = compute_asset_exposure(portfolio, funds_by_abbr)
    for slot_key, label in _ASSET_LABELS.items():
        # Labels carry &amp; (e.g. "Money Market &amp; Cash"); unescape to match
        # the rendered (also-&amp;-encoded) legend label after our own unescape.
        authoritative[_html.unescape(label)] = asset[slot_key]
    for geo_label, pct, _ in compute_geo_exposure(portfolio, funds_by_abbr):
        authoritative[_html.unescape(geo_label)] = pct

    sect_m = re.search(
        r'Portfolio Exposure</div>(.*?)(?=<div class="section">|</body>)',
        html_text,
        re.DOTALL,
    )
    if sect_m is None:
        return violations
    section = sect_m.group(1)

    pattern = re.compile(
        r'legend-label">([^<]*)</span>\s*<span class="legend-pct">([^<]*)<'
    )
    for label_raw, pct_raw in pattern.findall(section):
        label = _html.unescape(label_raw).strip()
        rendered = _fidelity_num(pct_raw)
        if rendered is None or label not in authoritative:
            continue
        if abs(rendered - authoritative[label]) > _FIDELITY_TOL:
            violations.append({
                "code": "RENDER_FIDELITY",
                "msg": (
                    f"Exposure '{label}': rendered {rendered}% != engine "
                    f"{authoritative[label]}%"
                ),
            })
    return violations


def _check_weighted_aggregate_fidelity(html_text: str, state: Any) -> list[dict[str, str]]:
    """Reconcile the weighted portfolio aggregates (weighted CFS, weighted 3Y alpha,
    weighted risk level) against _compute_portfolio_metrics (the production function)."""
    from consultant_engine.nodes.generate_proposal import (  # local: avoid cycle
        _compute_portfolio_metrics,
    )

    metrics = _compute_portfolio_metrics(
        state.get("portfolio", []),
        state.get("cfs_scores", []),
        state.get("eligible_funds", []),
    )
    # data-slot key → (authoritative value, human label).
    checks = {
        "portfolio.cfs_composite": (metrics["weighted_cfs"], "weighted CFS"),
        "portfolio.weighted_cfs": (metrics["weighted_cfs"], "weighted CFS"),
        "portfolio.alpha_3y": (metrics["weighted_alpha_3y"], "weighted 3Y alpha"),
        "portfolio.weighted_alpha": (metrics["weighted_alpha_3y"], "weighted 3Y alpha"),
        "portfolio.weighted_risk_level": (metrics["weighted_risk_level"], "weighted risk level"),
    }

    violations: list[dict[str, str]] = []
    seen: set[str] = set()
    for slot_key, (auth_str, label) in checks.items():
        auth = _fidelity_num(auth_str)
        if auth is None:
            continue
        for m in re.finditer(
            rf'data-slot="{re.escape(slot_key)}"[^>]*>([^<]*)<', html_text
        ):
            rendered = _fidelity_num(m.group(1))
            if rendered is None:
                continue
            if abs(rendered - auth) > _FIDELITY_TOL:
                msg = (
                    f"Portfolio {label} (slot {slot_key}): rendered {rendered} != "
                    f"engine {auth}"
                )
                if msg in seen:
                    continue
                seen.add(msg)
                violations.append({"code": "RENDER_FIDELITY", "msg": msg})
    return violations


def check_render_fidelity(html_text: str, state: Any) -> list[dict[str, str]]:
    """Reconcile the rendered Python-owned values in ``html_text`` against the
    engine's OWN authoritative values in ``state``.

    Runs in the runtime ``validate`` node (which holds state) — NOT in
    ``validate_html`` (the offline eval layer has no state). Defends against the
    repair pass, which round-trips the whole document through the LLM and so could
    silently corrupt a Python-owned number, label, or allocation.

    Covers a coverage canary (every portfolio holding renders as a fund card, so a
    dropped/renamed card can't make the per-card checks pass vacuously) and, per fund
    card: CFS composite, allocation %, Shariah label, Risk Level, Volatility Factor;
    the asset-class + geographic exposure legend percentages; and the weighted
    portfolio aggregates (weighted CFS, weighted 3Y alpha, weighted risk level).
    Authoritative values come from state or from REUSING the production compute
    functions — never re-derived here.

    Args:
        html_text: The final (post-generate or post-repair) proposal HTML.
        state: The ConsultantState (reads cfs_scores, portfolio, eligible_funds).

    Returns:
        A list of ``{"code": "RENDER_FIDELITY", "msg": ...}`` dicts; ``[]`` when every
        rendered Python-owned value matches its authoritative source.
    """
    return (
        _check_card_coverage(html_text, state)
        + _check_fund_card_fidelity(html_text, state)
        + _check_exposure_fidelity(html_text, state)
        + _check_weighted_aggregate_fidelity(html_text, state)
    )


# ── Composite runner ──────────────────────────────────────────────────────────


def validate_html(
    html_text: str,
    version: str,
    wb_index: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """Run all validation rules and return the concatenated list of violations.

    Returns ``[]`` for a clean proposal.
    """
    return (
        check_sections(html_text)
        + check_version_and_disclosure(html_text, version)
        + check_cfs_consistency(html_text)
        + check_perf_consistency(html_text)
        + check_exposure_consistency(html_text)
        + check_summary_consistency(html_text)
        + check_funds_in_workbook(html_text, wb_index)
        + check_alpha_warning(html_text, wb_index)
        + check_retail_eligibility(html_text, wb_index)
        + check_unfilled_slots(html_text)
    )
