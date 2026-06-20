"""load_funds node: read the FundMaster workbook into per-fund records.

Parses the 'Master' sheet of a screened FundMaster .xlsx into the Fund shape used
downstream, dropping rows that are not retail-eligible (B-series and wholesale
classes). This is the pipeline's bridge from the screener's Excel output to the
consultant engine's in-memory state.
"""

import openpyxl
from consultant_engine.state import ConsultantState

# Step 1b: funds excluded from retail eligibility
WHOLESALE = {"PBCPF", "PWSIF", "PIWSIF", "PeWS20F"}


def _excluded(name: str, abbr: str) -> bool:
    """True if a fund row is not retail-eligible and should be skipped.

    Drops the "PB "-named B-series share classes (duplicate the A-series holdings),
    any "-B" abbreviation, and the named WHOLESALE funds. Missing name/abbr is
    treated as excluded.
    """
    if name is None or abbr is None:
        return True
    return name.startswith("PB ") or abbr.endswith("-B") or abbr in WHOLESALE


def _f(value):
    """Excel cell → float, preserving None (a missing value is not 0.0)."""
    return float(value) if value is not None else None


def _s(value):
    """Excel cell → stripped str, preserving None."""
    return str(value).strip() if value is not None else None


def _read_cols(ws, row, start_col, keys) -> dict:
    """Read len(keys) consecutive cells from start_col into {key: float-or-None}."""
    return {key: _f(ws.cell(row, start_col + i).value) for i, key in enumerate(keys)}


def load_funds(state: ConsultantState) -> dict:
    """load_funds node: parse the FundMaster 'Master' sheet into Fund records.

    Reads state["fundmaster_path"], iterates the data rows until the abbr column is
    empty, skips retail-ineligible rows (see _excluded), and builds the typed Fund
    dict for each survivor. Returns {"eligible_funds": [...]} — the retail superset
    before Shariah / risk-ceiling filtering.
    """
    path = state["fundmaster_path"]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master"]

    # Read geo header names from row 3, cols 41–52
    geo_headers = []
    for col in range(41, 53):
        cell_val = ws.cell(3, col).value
        geo_headers.append(cell_val if cell_val is not None else f"col{col}")

    # Period names in order, starting at col 15; 5 periods × 3 cols each
    PERIOD_NAMES = ["ytd", "1y", "3y", "5y", "10y"]
    ALPHA_EFFICIENCY_KEYS = ["ytd", "1y", "3y", "5y", "10y"]
    ASSET_KEYS = ["dom_equity", "for_equity", "fi", "mm", "deposits", "other"]

    eligible_funds = []

    # Data starts at row 4; stop when col 2 (abbr) is empty
    row = 4
    while True:
        abbr = ws.cell(row, 2).value
        if abbr is None or str(abbr).strip() == "":
            break

        abbr = str(abbr).strip()
        name_val = ws.cell(row, 1).value
        name = str(name_val).strip() if name_val is not None else ""

        if _excluded(name, abbr):
            row += 1
            continue

        shariah_raw = ws.cell(row, 3).value
        shariah = True if str(shariah_raw).strip().lower() == "yes" else False

        fund_type = ws.cell(row, 4).value
        risk_level_raw = ws.cell(row, 6).value
        risk_level = int(risk_level_raw) if risk_level_raw is not None else None

        status = _s(ws.cell(row, 10).value)
        weighted_alpha = _f(ws.cell(row, 14).value)

        # Returns: cols 15–29, 5 periods × 3 cols each (fund/bench/alpha)
        returns = {}
        for i, period in enumerate(PERIOD_NAMES):
            base_col = 15 + i * 3
            returns[period] = {
                "fund": _f(ws.cell(row, base_col).value),
                "bench": _f(ws.cell(row, base_col + 1).value),
                "alpha": _f(ws.cell(row, base_col + 2).value),
            }

        alpha_efficiency = _read_cols(ws, row, 30, ALPHA_EFFICIENCY_KEYS)  # cols 30–34
        assets = _read_cols(ws, row, 35, ASSET_KEYS)                       # cols 35–40
        geo = _read_cols(ws, row, 41, geo_headers)                         # cols 41–52, headers from row 3

        # Top-5 holdings — col 64 is a " | "-delimited string in the real workbook
        # (build_sheet_data.py joins with ' | '). Split to a list so overlap checks
        # compare holding names, not characters.
        top5_raw = ws.cell(row, 64).value
        if top5_raw is None or top5_raw == "":
            top5_holdings = []
        else:
            top5_holdings = [h.strip() for h in str(top5_raw).split("|") if h.strip()]

        volatility_factor = _f(ws.cell(row, 65).value)
        lipper_class = _s(ws.cell(row, 67).value)
        benchmark = _s(ws.cell(row, 68).value)
        drawdown = _f(ws.cell(row, 72).value)  # 0.0 is a valid drawdown — preserved by _f

        days_raw = ws.cell(row, 73).value
        days_from_ath = int(days_raw) if days_raw is not None else None

        fund: dict = {
            "abbr": abbr,
            "name": name,
            "shariah": shariah,
            "fund_type": fund_type,
            "risk_level": risk_level,
            "status": status,
            "weighted_alpha": weighted_alpha,
            "returns": returns,
            "alpha_efficiency": alpha_efficiency,
            "assets": assets,
            "geo": geo,
            "top5_holdings": top5_holdings,
            "volatility_factor": volatility_factor,
            "lipper_class": lipper_class,
            "benchmark": benchmark,
            "drawdown": drawdown,
            "days_from_ath": days_from_ath,
        }
        eligible_funds.append(fund)
        row += 1

    wb.close()
    return {"eligible_funds": eligible_funds}
