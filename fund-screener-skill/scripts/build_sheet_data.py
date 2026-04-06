"""
Build clean structured CSV from mfr_results.json (v4 — enhanced Master sheet)
New in v4:
  - Fund Type from PHS 'Category of Fund' (Equity, Mixed Asset / Balanced, Fixed Income, Money Market, Fund of Funds)
  - Status moved to SCREENING band
  - Alpha Efficiency columns (Alpha / VF) for YTD, 1Y, 3Y, 5Y, 10Y
  - Sector breakdown into named columns (like geo breakdown)
  - Risk Level from MFR TOC
Outputs: master_funds.csv (all funds, qualified + disqualified)
"""

import json
import csv
import os
import sys
import openpyxl

# ── Paths (auto-derived from script location: scripts/ → fund-screener-skill/ → Funds/) ──
WORK_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# RISK_LEVEL_FILE: leave as "" to auto-detect from the Funds folder
# (the script looks for funds_risk_level.xlsx one level above BASE_DIR).
# Set an explicit path only if the file lives somewhere non-standard.
RISK_LEVEL_FILE = ""

if not WORK_DIR:
    print("ERROR: Please set WORK_DIR at the top of this script.")
    sys.exit(1)

with open(os.path.join(WORK_DIR, "mfr_results.json")) as f:
    data = json.load(f)

all_funds_raw = data['all_funds']
qualified_raw = data['qualified']

# ── Auto-detect risk file from Funds folder (parent of BASE_DIR) ─────────────
if not RISK_LEVEL_FILE:
    base_dir_from_json = data.get('base_dir', '')
    if base_dir_from_json:
        # BASE_DIR is the "Unit Trust (UT)" subfolder; Funds folder is one level up
        funds_dir = os.path.dirname(base_dir_from_json)
        candidate = os.path.join(funds_dir, "funds_risk_level.xlsx")
        if os.path.exists(candidate):
            RISK_LEVEL_FILE = candidate
            print(f"Auto-detected risk level file: {RISK_LEVEL_FILE}")
        else:
            print(f"WARNING: funds_risk_level.xlsx not found in {funds_dir}")

# ── Load Risk Levels from Excel ──────────────────────────────────────────────
risk_map = {}
if RISK_LEVEL_FILE and os.path.exists(RISK_LEVEL_FILE):
    try:
        rl_wb = openpyxl.load_workbook(RISK_LEVEL_FILE)
        rl_ws = rl_wb.active
        for r in range(2, rl_ws.max_row + 1):
            abbr = rl_ws.cell(row=r, column=2).value
            rl = rl_ws.cell(row=r, column=3).value
            if abbr and rl is not None:
                risk_map[abbr] = int(rl)
                risk_map[abbr.replace(' ', '').upper()] = int(rl)
        print(f"Loaded {len(set(risk_map.values()))} unique risk levels for {rl_ws.max_row - 1} funds")
    except Exception as e:
        print(f"WARNING: Could not load risk levels: {e}")
else:
    print("WARNING: funds_risk_level.xlsx not found — Risk Level column will be empty")


def standardize_distribution(raw):
    """Standardize distribution policy to: Annual, Monthly, Semi-Annual, Incidental, None."""
    if not raw:
        return 'None'
    low = raw.lower()
    if 'no distribution' in low:
        return 'None'
    if 'semi-annual' in low or 'semi annual' in low:
        return 'Semi-Annual'
    if 'monthly' in low:
        return 'Monthly'
    if 'annual' in low:
        return 'Annual'
    if 'incidental' in low:
        return 'Incidental'
    return 'None'


qualified_abbrs = set(f['abbr'] for f in qualified_raw)

# ── Weighted Alpha calculation (v8) ─────────────────────────────────────────
# Recompute weighted alpha for all funds (extract_mfr.py stores it in fund_data,
# but we recompute here from period_detail to ensure consistency with CSV output)
WA_WEIGHTS = {'ytd': 0.05, '1-year': 0.15, '3-year': 0.40, '5-year': 0.25, '10-year': 0.15}

def compute_weighted_alpha(fund):
    """Compute weighted alpha score from period_detail alphas."""
    pd = fund.get('period_detail', {})
    available = {}
    for period_key, weight in WA_WEIGHTS.items():
        entry = pd.get(period_key, {})
        alpha = entry.get('alpha')
        if alpha is not None and alpha != '':
            try:
                available[period_key] = (weight, float(alpha))
            except (ValueError, TypeError):
                pass
    # Also check ytd via the separate ytd field if period_detail doesn't have it
    if 'ytd' not in available:
        ytd_old = fund.get('ytd') or {}
        ytd_f = ytd_old.get('fund', '')
        ytd_b = ytd_old.get('benchmark', '')
        if ytd_f != '' and ytd_b != '':
            try:
                alpha = round(float(ytd_f) - float(ytd_b), 2)
                available['ytd'] = (WA_WEIGHTS['ytd'], alpha)
            except (ValueError, TypeError):
                pass
    if len(available) < 2:
        return None
    total_w = sum(w for w, _ in available.values())
    return round(sum((w / total_w) * a for w, a in available.values()), 4)

# Rebuild qualified_abbrs using weighted alpha (>0)
qualified_abbrs = set()
for f in all_funds_raw:
    wa = compute_weighted_alpha(f)
    f['_weighted_alpha'] = wa
    tp = f.get('total_periods', 0)
    if wa is not None and wa > 0 and tp >= 2:
        qualified_abbrs.add(f['abbr'])


def build_rationale(fund, status, out_p, total_p):
    """Build a concise text rationale summarising which periods beat/missed the benchmark."""
    period_detail = fund.get('period_detail', {})

    # YTD alpha — mirror the same fallback logic used in build_row()
    ytd_d = period_detail.get('ytd', {})
    ytd_f = ytd_d.get('fund', '')
    ytd_b = ytd_d.get('benchmark', '')
    ytd_a = ytd_d.get('alpha', '')
    if ytd_f == '':
        ytd_old = fund.get('ytd') or {}
        ytd_f = ytd_old.get('fund', '')
        ytd_b = ytd_old.get('benchmark', '')
    if ytd_a == '' and ytd_f != '' and ytd_b != '':
        try:
            ytd_a = round(float(ytd_f) - float(ytd_b), 2)
        except (ValueError, TypeError):
            pass

    checks = [
        ('YTD', ytd_a),
        ('1Y',  period_detail.get('1-year',  {}).get('alpha', '')),
        ('3Y',  period_detail.get('3-year',  {}).get('alpha', '')),
        ('5Y',  period_detail.get('5-year',  {}).get('alpha', '')),
        ('10Y', period_detail.get('10-year', {}).get('alpha', '')),
    ]

    parts = []
    for label, alpha in checks:
        if alpha != '' and alpha is not None:
            try:
                parts.append(f"{label}{'✔' if float(alpha) > 0 else '✘'}")
            except (ValueError, TypeError):
                pass

    periods_str = ' '.join(parts) if parts else 'N/A'
    prefix = 'Qualified' if status == 'Qualified' else 'Disqualified'
    wa = fund.get('_weighted_alpha')
    wa_str = f"WA: {wa:+.2f}%" if wa is not None else "WA: N/A"
    return f"{prefix} ({wa_str}) — {periods_str}"


def clean_name(name):
    return name.replace('EPF Qualified Fund', '').replace('EPFQualifiedFund', '').replace('Fund Award Won', '').strip()


def fmt(val):
    if val is None or val == '':
        return ''
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


def get_period(fund, period):
    d = fund.get('period_detail', {})
    if period in d:
        entry = d[period]
        return entry.get('fund', ''), entry.get('benchmark', ''), entry.get('alpha', '')
    return '', '', ''


def split_allocation(alloc):
    """Map raw MFR allocation keys into 6 consolidated columns."""
    ALLOC_MAP = {
        'Equity & equity-related securities - Domestic': 'Domestic Equity',
        'Shariah-compliant equity & Shariah-compliant equity-related securities - Domestic': 'Domestic Equity',
        '- Domestic Shariah-compliant equity & Shariah-compliant equity-related securities': 'Domestic Equity',
        'Equity & equity-related securities - Foreign': 'Foreign Equity',
        'Shariah-compliant equity & Shariah-compliant equity-related securities - Foreign': 'Foreign Equity',
        'Shariah-compliant equity & Shariah-compliant equity-related securities': 'Foreign Equity',
        'Fixed income securities': 'Fixed Income / Sukuk',
        'Sukuk': 'Fixed Income / Sukuk',
        '- Foreign Sukuk': 'Fixed Income / Sukuk',
        'Money market instruments & others': 'Money Market',
        'Islamic money market instruments & others': 'Money Market',
        '- Foreign Islamic money market instruments & others': 'Money Market',
        'Deposits and money market instruments': 'Deposits',
        'Islamic deposits and Islamic money market instruments': 'Deposits',
    }
    result = {}
    for raw_key, pct in (alloc or {}).items():
        col = ALLOC_MAP.get(raw_key, 'Other')
        result[col] = round(result.get(col, 0) + pct, 1)
    return result


ALLOC_COLUMNS = ['Domestic Equity', 'Foreign Equity', 'Fixed Income / Sukuk',
                 'Money Market', 'Deposits', 'Other']

GEO_COLUMNS = ['USA', 'Taiwan', 'Korea', 'Japan', 'France', 'Germany', 'China',
               'Singapore', 'Netherlands', 'Indonesia', 'Australia', 'Geo Other']

# ── Sector breakdown columns ─────────────────────────────────────────────────
SECTOR_MAP = {
    'Industrial': 'Industrial',
    'Technology': 'Technology',
    'Financial': 'Financial',
    'Communications': 'Communications',
    'Telecommunications': 'Communications',
    'Internet': 'Communications',
    'Consumer, Cyclical': 'Consumer Discretionary',
    'Consumer, Non-cyclical': 'Consumer Staples',
    'Utilities': 'Utilities',
    'Energy': 'Energy',
    'Electric': 'Energy',
    'Basic Materials': 'Materials',
    'Real Estate': 'Real Estate',
    'REITS': 'Real Estate',
    'Home Builders': 'Real Estate',
    'ETF': 'Other Sector',
    'Diversified': 'Other Sector',
}

SECTOR_COLUMNS = ['Industrial', 'Technology', 'Financial', 'Communications',
                  'Consumer Discretionary', 'Consumer Staples', 'Utilities',
                  'Energy', 'Materials', 'Real Estate', 'Other Sector']


def split_sectors(top5_sectors):
    """Map top 5 sector entries into named sector columns (aggregated %)."""
    result = {}
    for s in (top5_sectors or []):
        name = s.get('sector', '')
        weight_str = s.get('weight', '0%').replace('%', '')
        try:
            weight = float(weight_str)
        except ValueError:
            weight = 0.0
        col = SECTOR_MAP.get(name, 'Other Sector')
        result[col] = round(result.get(col, 0) + weight, 1)
    return result


def split_geo(geo):
    """Map geo breakdown into named country columns + Other bucket."""
    NAMED = {'USA', 'Taiwan', 'Korea', 'Japan', 'France', 'Germany', 'China',
             'Singapore', 'Netherlands', 'Indonesia', 'Australia'}
    result = {}
    other_total = 0.0
    for country, pct in (geo or {}).items():
        if country in NAMED:
            result[country] = pct
        else:
            other_total += pct
    if other_total > 0:
        result['Geo Other'] = round(other_total, 1)
    return result


def build_row(fund):
    name = clean_name(fund['name'])
    abbr = fund['abbr']
    rate = fund['outperform_rate']
    total_p = fund['total_periods']
    out_p = fund['outperform_count']

    # Status
    status = 'Qualified' if abbr in qualified_abbrs else 'Disqualified'

    # Fund Type from PHS Category (primary) or fallback to derived asset_class
    fund_type = fund.get('phs_fund_type', '')
    if not fund_type:
        # Fallback: derive from old asset_class
        ac = fund.get('asset_class', fund.get('fund_type', ''))
        if 'Money' in ac:
            fund_type = 'Money Market'
        elif any(x in ac for x in ['Bond', 'Sukuk', 'Fixed']):
            fund_type = 'Fixed Income'
        elif any(x in ac for x in ['Mixed', 'Balanced']):
            fund_type = 'Mixed Asset / Balanced'
        elif 'Equity' in ac:
            fund_type = 'Equity'
        else:
            fund_type = 'Equity'

    # Risk Level lookup
    rl = risk_map.get(abbr, risk_map.get(abbr.replace(' ', '').upper(), ''))

    # VF for Alpha Efficiency
    vf = fund.get('volatility_factor', '')
    vf_float = None
    if vf:
        try:
            vf_float = float(vf)
        except (ValueError, TypeError):
            pass

    # YTD
    ytd_detail = fund.get('period_detail', {}).get('ytd', {})
    ytd_fund = ytd_detail.get('fund', '')
    ytd_bench = ytd_detail.get('benchmark', '')
    ytd_alpha = ytd_detail.get('alpha', '')

    if ytd_fund == '':
        ytd_old = fund.get('ytd') or {}
        ytd_fund = ytd_old.get('fund', '')
        ytd_bench = ytd_old.get('benchmark', '')
        if ytd_fund != '' and ytd_bench != '':
            try:
                ytd_alpha = round(float(ytd_fund) - float(ytd_bench), 2)
            except:
                ytd_alpha = ''

    # Period data
    f1, b1, a1 = get_period(fund, '1-year')
    f3, b3, a3 = get_period(fund, '3-year')
    f5, b5, a5 = get_period(fund, '5-year')
    f10, b10, a10 = get_period(fund, '10-year')

    # Alpha Efficiency = Alpha / VF
    def alpha_eff(alpha_val):
        if alpha_val == '' or alpha_val is None or vf_float is None or vf_float == 0:
            return ''
        try:
            return round(float(alpha_val) / vf_float, 2)
        except:
            return ''

    # Asset allocation & geo & sector splits
    alloc_split = split_allocation(fund.get('asset_allocation', {}))
    geo_split = split_geo(fund.get('geo_breakdown', {}))
    sector_split = split_sectors(fund.get('top5_sectors', []))

    row = {
        # ── FUND DETAILS ──
        'Fund Name': name,
        'Abbr': abbr,
        'Series': 'Shariah' if fund.get('is_shariah') else 'Conventional',
        'Fund Type': fund_type,
        'Geography': fund.get('geography', ''),
        'Objective': fund.get('objective_class', ''),
        'Risk Level': rl,
        'Distribution Policy': standardize_distribution(fund.get('distribution_policy', '')),
        'Fund Size (RM Mil)': fund.get('fund_size_rm_mil', ''),
        'Launch Date': fund.get('launch_date', ''),
        # ── SCREENING ──
        'Status': status,
        'Outperform Rate (%)': rate,
        'Periods Assessed': f"{out_p}/{total_p}",
        'Rationale': build_rationale(fund, status, out_p, total_p),
        'Weighted Alpha (%)': fmt(fund.get('_weighted_alpha', '')),
        # ── RETURNS ──
        'YTD Fund (%)': fmt(ytd_fund),
        'YTD Benchmark (%)': fmt(ytd_bench),
        'YTD Alpha': fmt(ytd_alpha),
        '1Y Fund (%)': fmt(f1),
        '1Y Benchmark (%)': fmt(b1),
        '1Y Alpha': fmt(a1),
        '3Y Fund (%)': fmt(f3),
        '3Y Benchmark (%)': fmt(b3),
        '3Y Alpha': fmt(a3),
        '5Y Fund (%)': fmt(f5),
        '5Y Benchmark (%)': fmt(b5),
        '5Y Alpha': fmt(a5),
        '10Y Fund (%)': fmt(f10),
        '10Y Benchmark (%)': fmt(b10),
        '10Y Alpha': fmt(a10),
        # ── ALPHA EFFICIENCY (Alpha / VF) ──
        'AE YTD': fmt(alpha_eff(ytd_alpha)),
        'AE 1Y': fmt(alpha_eff(a1)),
        'AE 3Y': fmt(alpha_eff(a3)),
        'AE 5Y': fmt(alpha_eff(a5)),
        'AE 10Y': fmt(alpha_eff(a10)),
    }
    # Asset Allocation columns
    for col in ALLOC_COLUMNS:
        row[f'Alloc: {col}'] = fmt(alloc_split.get(col, ''))
    # Geo Breakdown columns
    for col in GEO_COLUMNS:
        row[f'Geo: {col}'] = fmt(geo_split.get(col, ''))
    # Sector Breakdown columns
    for col in SECTOR_COLUMNS:
        row[f'Sector: {col}'] = fmt(sector_split.get(col, ''))
    # Top 5 text
    row['Top 5 Holdings'] = ' | '.join(fund.get('top5_holdings', []))
    row['Top 5 Sectors'] = ' | '.join([f"{s['sector']} ({s['weight']})" for s in fund.get('top5_sectors', [])])
    # Meta
    row['Volatility Factor'] = fund.get('volatility_factor', '')
    row['Volatility Class'] = fund.get('volatility_class', '')
    row['Lipper Class'] = fund.get('lipper_class', '')
    row['Benchmark'] = fund.get('benchmark', '')
    row['Source MFR'] = fund.get('source_file', '')
    return row


# Sort: qualified first, then by fund type, weighted alpha, 3Y alpha
def sort_key(r):
    type_order = {
        'Equity': 1,
        'Mixed Asset / Balanced': 2,
        'Fixed Income': 3,
        'Fund of Funds': 4,
        'Money Market': 5,
    }
    status_order = 0 if r['Status'] == 'Qualified' else 1
    ft = type_order.get(r['Fund Type'], 9)
    wa = -float(r['Weighted Alpha (%)']) if r['Weighted Alpha (%)'] else 0
    alpha3 = -float(r['3Y Alpha']) if r['3Y Alpha'] else 0
    return (status_order, ft, wa, alpha3)


# Build ALL fund rows
all_rows = [build_row(f) for f in all_funds_raw]
all_rows.sort(key=sort_key)

# Write master CSV
csv_path = os.path.join(WORK_DIR, "master_funds.csv")
fieldnames = list(all_rows[0].keys())
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(all_rows)

qual_count = sum(1 for r in all_rows if r['Status'] == 'Qualified')
disq_count = len(all_rows) - qual_count
rl_filled = sum(1 for r in all_rows if r.get('Risk Level', '') != '')

print(f"Written {len(all_rows)} funds to {csv_path}")
print(f"  Qualified: {qual_count} | Disqualified: {disq_count}")
print(f"  Risk Level coverage: {rl_filled}/{len(all_rows)}")

from collections import Counter
ft_counts = Counter(r['Fund Type'] for r in all_rows if r['Status'] == 'Qualified')
print("\nQualified funds by fund type:")
for ft, count in sorted(ft_counts.items(), key=lambda x: x[0]):
    print(f"  {ft}: {count}")

shariah_count = sum(1 for r in all_rows if r['Series'] == 'Shariah' and r['Status'] == 'Qualified')
print(f"\nShariah: {shariah_count} | Conventional: {qual_count - shariah_count}")
