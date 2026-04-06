"""
Build a professionally formatted Excel workbook for Public Mutual funds (v5).
Sheet 1: Master (all funds — qualified + disqualified)
Sheet 2: Summary dashboard matching PDF structure with MEDIAN stats + Top 5 per fund type

New in v5:
  - Top 5 Sectors column removed
  - All percentage columns (L-BJ) store raw % values with plain number formats
  - Percentage column headers updated with "(%) " suffix
  - Summary sheet uses formula-driven Top-5 tables with SORTBY/FILTER/INDEX
  - Median statistics use MEDIAN(FILTER()) formulas instead of Python hardcoding
  - Beat % color scale adjusted for 0-100 range instead of 0.0-1.0
"""

import csv
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PATHS                                                           ║
# ╚══════════════════════════════════════════════════════════════════╝
WORK_DIR = "/sessions/nifty-vibrant-cannon/mnt/Funds"
OUT_PATH = "/sessions/nifty-vibrant-cannon/mnt/Funds/PublicMutual_FundMaster_Feb2026_v6.xlsx"

if not WORK_DIR or not OUT_PATH:
    print("ERROR: Please set WORK_DIR and OUT_PATH at the top of this script.")
    sys.exit(1)

# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW   = "EBF3FB"
C_WHITE     = "FFFFFF"

def thin_border(color="BDD7EE"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def read_csv(path):
    with open(path, encoding='utf-8') as f:
        return list(csv.DictReader(f))

def flt(val):
    try:
        return float(val) if val not in ('', None) else None
    except:
        return None

# ── Load data ────────────────────────────────────────────────────────────────
rows = read_csv(os.path.join(WORK_DIR, "master_funds.csv"))
mfr_data = json.load(open(os.path.join(WORK_DIR, "mfr_results.json")))

# ── ATH data (v6) ─────────────────────────────────────────────────────────────
ath_path = os.path.join(WORK_DIR, "ath_results.json")
ath_map = {}
if os.path.exists(ath_path):
    ath_data = json.load(open(ath_path))
    for rec in ath_data.get("funds", []):
        ath_map[rec["abbr"]] = rec
    print(f"ATH data loaded: {len(ath_map)} funds")
else:
    print("WARNING: ath_results.json not found — ATH columns will be empty")

# Enrich rows with ATH fields
for row in rows:
    abbr = row.get("Abbr", "")
    ath = ath_map.get(abbr, {})
    row["_ath_nav"]     = ath.get("ath_nav", "")
    row["_ath_date"]    = ath.get("ath_date", "")
    row["_cur_nav"]     = ath.get("current_nav", "")
    row["_drawdown"]    = ath.get("drawdown_pct", "")
    row["_days_ath"]    = ath.get("days_from_ath", "")

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: MASTER
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Master"
ws.sheet_view.showGridLines = False

# Column definitions: (header, csv_key, width, number_format, is_formula)
# Layout v5:
#   FUND DETAILS(1-10): Fund Name, Abbr, Series, Fund Type, Geography, Objective, Risk Level, Distribution, Size, Launch
#   SCREENING(11-13): Status, Beat %, Periods
#   RETURNS(14-28): YTD/1Y/3Y/5Y/10Y × (Fund, Bench, Alpha) — all store raw % values
#   ALPHA EFFICIENCY(29-33): AE YTD, AE 1Y, AE 3Y, AE 5Y, AE 10Y (formula: Alpha/VF)
#   ASSET ALLOCATION(34-39): 6 cols — all store raw % values
#   GEO BREAKDOWN(40-51): 12 cols — all store raw % values
#   SECTOR BREAKDOWN(52-62): 11 cols — all store raw % values
#   TOP 5(63): Holdings text only (Top 5 Sectors removed)
#   META(64-67): VF, VC, Lipper Class, Benchmark

COLS = [
    # ── Fund Details (1-10) ─────────────
    ("Fund Name",        "Fund Name",           40, "@",       False),  # 1
    ("Abbr",             "Abbr",                12, "@",       False),  # 2
    ("Series",           "Series",              14, "@",       False),  # 3
    ("Fund Type",        "Fund Type",           22, "@",       False),  # 4
    ("Geography",        "Geography",           16, "@",       False),  # 5
    ("Objective",        "Objective",           24, "@",       False),  # 6
    ("Risk Level",       "Risk Level",          10, "0",       False),  # 7
    ("Distribution",     "Distribution Policy", 16, "@",       False),  # 8
    ("Size (RM M)",      "Fund Size (RM Mil)",  14, "#,##0.00",False),  # 9
    ("Launch",           "Launch Date",         12, "@",       False),  # 10
    # ── Screening (11-13) ────────────────
    ("Status",           "Status",              14, "@",       False),  # 11
    ("Beat (%)",         "Outperform Rate (%)", 10, "0.0",    False),  # 12
    ("Periods",          "Periods Assessed",    10, "@",       False),  # 13
    # ── Returns (14-28) — all raw % values ──
    ("YTD Fund (%)",     "YTD Fund (%)",        10, "0.00",   False),  # 14
    ("YTD Bench (%)",    "YTD Benchmark (%)",   10, "0.00",   False),  # 15
    ("YTD Alpha (%)",    "YTD Alpha",           10, "0.00",   False),  # 16
    ("1Y Fund (%)",      "1Y Fund (%)",         10, "0.00",   False),  # 17
    ("1Y Bench (%)",     "1Y Benchmark (%)",    10, "0.00",   False),  # 18
    ("1Y Alpha (%)",     "1Y Alpha",            10, "0.00",   False),  # 19
    ("3Y Fund (%)",      "3Y Fund (%)",         10, "0.00",   False),  # 20
    ("3Y Bench (%)",     "3Y Benchmark (%)",    10, "0.00",   False),  # 21
    ("3Y Alpha (%)",     "3Y Alpha",            10, "0.00",   False),  # 22
    ("5Y Fund (%)",      "5Y Fund (%)",         10, "0.00",   False),  # 23
    ("5Y Bench (%)",     "5Y Benchmark (%)",    10, "0.00",   False),  # 24
    ("5Y Alpha (%)",     "5Y Alpha",            10, "0.00",   False),  # 25
    ("10Y Fund (%)",     "10Y Fund (%)",        10, "0.00",   False),  # 26
    ("10Y Bench (%)",    "10Y Benchmark (%)",   10, "0.00",   False),  # 27
    ("10Y Alpha (%)",    "10Y Alpha",           10, "0.00",   False),  # 28
    # ── Alpha Efficiency (29-33) — formula columns ──
    ("AE YTD",           "AE_FORMULA",          9,  "0.00",   True),   # 29
    ("AE 1Y",            "AE_FORMULA",          9,  "0.00",   True),   # 30
    ("AE 3Y",            "AE_FORMULA",          9,  "0.00",   True),   # 31
    ("AE 5Y",            "AE_FORMULA",          9,  "0.00",   True),   # 32
    ("AE 10Y",           "AE_FORMULA",          9,  "0.00",   True),   # 33
    # ── Asset Allocation (34-39) — raw % values ─
    ("Dom. Equity (%)",  "Alloc: Domestic Equity",       11, "0.0",   False),  # 34
    ("For. Equity (%)",  "Alloc: Foreign Equity",        11, "0.0",   False),  # 35
    ("FI / Sukuk (%)",   "Alloc: Fixed Income / Sukuk",  11, "0.0",   False),  # 36
    ("Money Mkt (%)",    "Alloc: Money Market",          11, "0.0",   False),  # 37
    ("Deposits (%)",     "Alloc: Deposits",              10, "0.0",   False),  # 38
    ("Alloc Other (%)",  "Alloc: Other",                 10, "0.0",   False),  # 39
    # ── Geo Breakdown (40-51) — raw % values ────
    ("USA (%)",          "Geo: USA",            8, "0.0",   False),   # 40
    ("Taiwan (%)",       "Geo: Taiwan",         8, "0.0",   False),   # 41
    ("Korea (%)",        "Geo: Korea",          8, "0.0",   False),   # 42
    ("Japan (%)",        "Geo: Japan",          8, "0.0",   False),   # 43
    ("France (%)",       "Geo: France",         8, "0.0",   False),   # 44
    ("Germany (%)",      "Geo: Germany",        8, "0.0",   False),   # 45
    ("China (%)",        "Geo: China",          8, "0.0",   False),   # 46
    ("Singapore (%)",    "Geo: Singapore",      9, "0.0",   False),   # 47
    ("Netherlands (%)",  "Geo: Netherlands",    10,"0.0",   False),   # 48
    ("Indonesia (%)",    "Geo: Indonesia",      9, "0.0",   False),   # 49
    ("Australia (%)",    "Geo: Australia",      9, "0.0",   False),   # 50
    ("Geo Other (%)",    "Geo: Geo Other",      9, "0.0",   False),   # 51
    # ── Sector Breakdown (52-62) — raw % values ─
    ("Industrial (%)",         "Sector: Industrial",          10, "0.0",   False),  # 52
    ("Technology (%)",         "Sector: Technology",           10, "0.0",   False),  # 53
    ("Financial (%)",          "Sector: Financial",            10, "0.0",   False),  # 54
    ("Comms (%)",              "Sector: Communications",       10, "0.0",   False),  # 55
    ("Cons. Disc. (%)",        "Sector: Consumer Discretionary",10,"0.0",   False),  # 56
    ("Cons. Staples (%)",      "Sector: Consumer Staples",     11, "0.0",   False),  # 57
    ("Utilities (%)",          "Sector: Utilities",            9,  "0.0",   False),  # 58
    ("Energy (%)",             "Sector: Energy",               8,  "0.0",   False),  # 59
    ("Materials (%)",          "Sector: Materials",             9,  "0.0",   False),  # 60
    ("Real Estate (%)",        "Sector: Real Estate",          10, "0.0",   False),  # 61
    ("Sector Other (%)",       "Sector: Other Sector",         10, "0.0",   False),  # 62
    # ── Top 5 (63) ──────────────────────────────
    ("Top 5 Holdings",   "Top 5 Holdings",      55, "@",   False),  # 63
    # ── Meta (64-67) ────────────────────────────
    ("VF",               "Volatility Factor",   6,  "0.0", False),   # 64
    ("VC",               "Volatility Class",    12, "@",   False),   # 65
    ("Lipper Class",     "Lipper Class",        28, "@",   False),   # 66
    ("Benchmark",        "Benchmark",           45, "@",   False),   # 67
    # ── ATH Momentum (68-72) — v6 ───────────────────────────────────
    ("ATH NAV",          "_ath_nav",            10, "0.0000", False), # 68
    ("ATH Date",         "_ath_date",           12, "@",      False), # 69
    ("Cur NAV",          "_cur_nav",            10, "0.0000", False), # 70
    ("Drawdown (%)",     "_drawdown",           11, "0.00",   False), # 71
    ("Days from ATH",    "_days_ath",           13, "0",      False), # 72
]

NUM_COLS = len(COLS)

# Column indices (1-based) for formula references
COL_VF = 64       # VF column (shifted from 65)
COL_YTD_ALPHA = 16
COL_1Y_ALPHA = 19
COL_3Y_ALPHA = 22
COL_5Y_ALPHA = 25
COL_10Y_ALPHA = 28
COL_AE_YTD = 29
COL_AE_1Y = 30
COL_AE_3Y = 31
COL_AE_5Y = 32
COL_AE_10Y = 33

# Map AE column index to its source Alpha column
AE_ALPHA_MAP = {
    COL_AE_YTD: COL_YTD_ALPHA,
    COL_AE_1Y:  COL_1Y_ALPHA,
    COL_AE_3Y:  COL_3Y_ALPHA,
    COL_AE_5Y:  COL_5Y_ALPHA,
    COL_AE_10Y: COL_10Y_ALPHA,
}

# ── Row 1: Title ─────────────────────────────────────────────────────────────
total_screened = len(mfr_data['all_funds'])
total_qual = sum(1 for r in rows if r.get("Status") == "Qualified")
title_text = (f"PUBLIC MUTUAL — FUND MASTER  |  "
              f"Data: MFR February 2026  |  "
              f"{total_qual} Qualified / {total_screened} Total Funds")
title_cell = ws.cell(row=1, column=1, value=title_text)
title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="1F3864")
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
ws.row_dimensions[1].height = 28

# ── Row 2: Group header bands ────────────────────────────────────────────────
GROUP_BANDS = [
    (1,  10, "FUND DETAILS",     "404040"),
    (11, 13, "SCREENING",        "C00000"),
    (14, 28, "ANNUALISED RETURNS vs BENCHMARK (MFR FEB 2026)", "2E75B6"),
    (29, 33, "ALPHA EFFICIENCY (Alpha / VF)", "1A5276"),
    (34, 39, "ASSET ALLOCATION (%)", "833C11"),
    (40, 51, "GEOGRAPHICAL BREAKDOWN (%)", "375623"),
    (52, 62, "SECTOR BREAKDOWN (%)", "2E4053"),
    (63, 63, "TOP 5",            "7030A0"),
    (64, 67, "META",             "808080"),
    (68, 72, "ATH MOMENTUM",    "1F618D"),
]

for start, end, label, color in GROUP_BANDS:
    cell = ws.cell(row=2, column=start, value=label)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border("FFFFFF")
    if end > start:
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    for c in range(start + 1, end + 1):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=2, column=c).border = thin_border("FFFFFF")

ws.row_dimensions[2].height = 22

# ── Row 3: Column headers ────────────────────────────────────────────────────
for col_idx, (hdr, key, width, nfmt, is_formula) in enumerate(COLS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=hdr)
    cell.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 36
ws.freeze_panes = "A4"

# ── Data rows ────────────────────────────────────────────────────────────────
data_start = 4
text_cols = {1, 2, 3, 4, 5, 6, 8, 10, 11, 13, 63, 65, 66, 67, 69}  # 69=ATH Date
wrap_cols = {6, 8, 63, 66, 67}
COL_DRAWDOWN = 71  # Drawdown % column index

C_DISQ_ROW = "F2F2F2"
C_DISQ_ROW_ALT = "E8E8E8"

for row_idx, fund in enumerate(rows, start=data_start):
    is_shariah = fund.get("Series", "") == "Shariah"
    is_disq = fund.get("Status", "") == "Disqualified"
    alt = (row_idx % 2 == 0)

    if is_disq:
        bg = C_DISQ_ROW_ALT if alt else C_DISQ_ROW
    else:
        bg = C_ALT_ROW if alt else C_WHITE

    for col_idx, (hdr, key, width, nfmt, is_formula) in enumerate(COLS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)

        if is_formula and col_idx in AE_ALPHA_MAP:
            # Alpha Efficiency formula: =IF(OR(Alpha="",VF="",VF=0),"",Alpha/VF)
            alpha_col_letter = get_column_letter(AE_ALPHA_MAP[col_idx])
            vf_col_letter = get_column_letter(COL_VF)
            cell.value = (f'=IF(OR({alpha_col_letter}{row_idx}="",'
                          f'{vf_col_letter}{row_idx}="",'
                          f'{vf_col_letter}{row_idx}=0),"",'
                          f'{alpha_col_letter}{row_idx}/{vf_col_letter}{row_idx})')
        else:
            raw = fund.get(key, "")
            # v5: all numeric columns use plain formats; store raw float directly
            if nfmt in ("#,##0.00", "0.00", "0.0", "0") and raw not in ("", None):
                try:
                    cell.value = float(raw)
                except:
                    cell.value = raw
            else:
                cell.value = raw if raw != "" else None

        cell.number_format = nfmt
        font_color = "999999" if is_disq else "000000"
        cell.font = Font(name="Arial", size=9, color=font_color,
                         italic=(is_shariah and col_idx <= 5))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(col_idx in wrap_cols),
            horizontal="left" if col_idx in text_cols else "center"
        )
        cell.border = thin_border()

    ws.row_dimensions[row_idx].height = 16

last_row = data_start + len(rows) - 1

# ── Conditional formatting ──────────────────────────────────────────────────

# Status column (11): green/red
status_col = get_column_letter(11)
ws.conditional_formatting.add(
    f"{status_col}{data_start}:{status_col}{last_row}",
    CellIsRule(operator='equal', formula=['"Qualified"'],
               fill=PatternFill("solid", fgColor="C6EFCE"),
               font=Font(name="Arial", size=9, color="276221")))
ws.conditional_formatting.add(
    f"{status_col}{data_start}:{status_col}{last_row}",
    CellIsRule(operator='equal', formula=['"Disqualified"'],
               fill=PatternFill("solid", fgColor="FCE4D6"),
               font=Font(name="Arial", size=9, color="9C0006")))

# Risk Level (col 7): color scale
rl_col = get_column_letter(7)
ws.conditional_formatting.add(
    f"{rl_col}{data_start}:{rl_col}{last_row}",
    ColorScaleRule(start_type='num', start_value=1, start_color="C6EFCE",
                   mid_type='num', mid_value=3, mid_color="FFEB9C",
                   end_type='num', end_value=5, end_color="FCE4D6"))

# Alpha columns: green positive, red negative
alpha_cols = [COL_YTD_ALPHA, COL_1Y_ALPHA, COL_3Y_ALPHA, COL_5Y_ALPHA, COL_10Y_ALPHA]
for col_idx in alpha_cols:
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="C6EFCE"),
                   font=Font(name="Arial", size=9, color="276221")))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="FCE4D6"),
                   font=Font(name="Arial", size=9, color="9C0006")))

# Alpha Efficiency: green positive, red negative
for col_idx in range(COL_AE_YTD, COL_AE_10Y + 1):
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="C6EFCE"),
                   font=Font(name="Arial", size=9, color="276221")))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="FCE4D6"),
                   font=Font(name="Arial", size=9, color="9C0006")))

# Beat % color scale (col 12) — now 0-100 instead of 0.0-1.0
beat_col = get_column_letter(12)
ws.conditional_formatting.add(
    f"{beat_col}{data_start}:{beat_col}{last_row}",
    ColorScaleRule(start_type='num', start_value=0, start_color="FCE4D6",
                   mid_type='num', mid_value=60, mid_color="FFEB9C",
                   end_type='num', end_value=100, end_color="C6EFCE"))

# Allocation columns — light bar (34-39)
for col_idx in range(34, 40):
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="FFF2CC")))

# Geo columns — light bar (40-51)
for col_idx in range(40, 52):
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="E2EFDA")))

# Sector columns — light bar (52-62)
for col_idx in range(52, 63):
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill("solid", fgColor="DAEEF3")))

# Drawdown % column (71): red scale — more negative = deeper red
dd_col = get_column_letter(COL_DRAWDOWN)
dd_rng = f"{dd_col}{data_start}:{dd_col}{last_row}"
ws.conditional_formatting.add(dd_rng,
    ColorScaleRule(start_type='num', start_value=-60, start_color="FCE4D6",
                   mid_type='num',   mid_value=-10,   mid_color="FFEB9C",
                   end_type='num',   end_value=0,     end_color="C6EFCE"))

# ── Auto-filter on header row ────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:{get_column_letter(NUM_COLS)}{last_row}"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: SUMMARY DASHBOARD (formula-driven, matching PDF structure)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False

SUMMARY_COLS = 8  # columns A-H

# ── Summary helper functions ─────────────────────────────────────────────────

# Master sheet reference prefix and column letters for formulas
MASTER_REF = "Master!"
M_STATUS = get_column_letter(11)     # Status col
M_FUND_TYPE = get_column_letter(4)   # Fund Type col
M_SERIES = get_column_letter(3)      # Series col
M_BEAT = get_column_letter(12)       # Beat %
M_NAME = get_column_letter(1)        # Fund Name
M_ABBR = get_column_letter(2)        # Abbr
M_3Y_FUND = get_column_letter(20)    # 3Y Fund
M_3Y_ALPHA = get_column_letter(22)   # 3Y Alpha
M_AE_3Y = get_column_letter(31)      # AE 3Y
M_SIZE = get_column_letter(9)        # Size
M_VC = get_column_letter(65)         # VC (shifted from 66)
M_YTD_FUND = get_column_letter(14)   # YTD Fund
M_1Y_FUND = get_column_letter(17)    # 1Y Fund
M_5Y_FUND = get_column_letter(23)    # 5Y Fund
M_10Y_FUND = get_column_letter(26)   # 10Y Fund

# Color palette for fund types
CLR_SCREENING = "C00000"
CLR_BREAKDOWN = "2E75B6"
CLR_EQUITY = "375623"
CLR_MIXED = "833C11"
CLR_FI = "7030A0"
CLR_MM = "2E75B6"
CLR_FOF = "1F3864"

def write_section_title(ws, row, text, color, cols=SUMMARY_COLS):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin_border("FFFFFF")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    for ci in range(2, cols + 1):
        ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row, column=ci).border = thin_border("FFFFFF")
    ws.row_dimensions[row].height = 26

def write_subhdr(ws, row, labels, bg="2E75B6"):
    for ci, lbl in enumerate(labels, start=1):
        c = ws.cell(row=row, column=ci, value=lbl)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 28

def write_data_row_vals(ws, row, vals, alt=False, bold=False, number_formats=None):
    bg = C_ALT_ROW if alt else C_WHITE
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(name="Arial", size=9, bold=bold)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="right" if ci > 2 else "left",
                                vertical="center", indent=1 if ci <= 2 else 0)
        if number_formats and ci <= len(number_formats) and number_formats[ci-1]:
            c.number_format = number_formats[ci-1]
    ws.row_dimensions[row].height = 16

def write_formula_row(ws, row, formulas, alt=False, number_formats=None):
    """Write a row of formulas (string starting with =) or plain values."""
    bg = C_ALT_ROW if alt else C_WHITE
    for ci, v in enumerate(formulas, start=1):
        c = ws.cell(row=row, column=ci)
        c.value = v
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="right" if ci > 2 else "left",
                                vertical="center", indent=1 if ci <= 2 else 0)
        if number_formats and ci <= len(number_formats) and number_formats[ci-1]:
            c.number_format = number_formats[ci-1]
    ws.row_dimensions[row].height = 16

def write_total_row(ws, row, vals, number_formats=None):
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")
        if number_formats and ci <= len(number_formats) and number_formats[ci-1]:
            c.number_format = number_formats[ci-1]
    ws.row_dimensions[row].height = 20


# Column widths for Summary
for ci, w in enumerate([32, 14, 12, 12, 12, 12, 14, 14], start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ── Title row ────────────────────────────────────────────────────────────────
r = 1
t = ws2.cell(row=r, column=1, value="PUBLIC MUTUAL FUND SCREENER — SUMMARY  |  MFR February 2026")
t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="1F3864")
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.merge_cells(f"A1:{get_column_letter(SUMMARY_COLS)}1")
for ci in range(2, SUMMARY_COLS + 1):
    ws2.cell(row=1, column=ci).fill = PatternFill("solid", fgColor="1F3864")
ws2.row_dimensions[1].height = 32

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1: SCREENING RESULTS (hardcoded stats)
# ═══════════════════════════════════════════════════════════════════════════════
r = 3
write_section_title(ws2, r, "SCREENING RESULTS", CLR_SCREENING)
r += 1
write_subhdr(ws2, r, ["Metric", "Value", "", "", "", "", "", ""], CLR_SCREENING)
r += 1

# Count PHS files
phs_total = 181
class_b = "PeCDF-B, PMMF-B, PBCMF-B, PeICDF-B, PIMMF-B, PBICMF-B"
wholesale = "PBCPF, PWSIF, PIWSIF, PeWS20F"
n_screened = total_screened

qual_rows = [x for x in rows if x.get("Status") == "Qualified"]
disq_rows_list = [x for x in rows if x.get("Status") == "Disqualified"]
n_shariah = sum(1 for x in qual_rows if x["Series"] == "Shariah")
n_conv = len(qual_rows) - n_shariah
pass_rate_val = len(qual_rows) / n_screened if n_screened else 0

stats = [
    ("Total PHS files (all UT funds)", phs_total, ""),
    ("Excluded: Class-B funds", f"6", f"({class_b})"),
    ("Excluded: Wholesale funds", f"4", f"({wholesale})"),
    ("Retail funds screened", n_screened, ""),
    (None, None, None),  # spacer
    ("Funds qualified (≥60% beat rate)", len(qual_rows), ""),
    ("Funds disqualified", len(disq_rows_list), ""),
    ("Pass rate", f"{pass_rate_val*100:.1f}%", ""),
    ("Conventional funds qualified", n_conv, ""),
    ("Shariah-compliant funds qualified", n_shariah, ""),
    (None, None, None),  # spacer
    ("Data source", "MFR February 2026", ""),
    ("Screening criteria", "≥60% of YTD/1Y/3Y/5Y/10Y periods outperform benchmark", ""),
    ("Minimum data requirement", "≥2 periods (YTD, 1Y, 3Y, 5Y, 10Y)", ""),
    ("Note", "Class-B & Wholesale funds excluded — not available to typical retail investors", ""),
]
for item in stats:
    if item[0] is None:
        r += 1
        continue
    vals = [item[0], item[1]]
    if item[2]:
        ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=SUMMARY_COLS)
        write_data_row_vals(ws2, r, vals, alt=(r % 2 == 0))
        c = ws2.cell(row=r, column=3, value=item[2])
        c.font = Font(name="Arial", size=8, italic=True, color="808080")
    else:
        write_data_row_vals(ws2, r, vals, alt=(r % 2 == 0))
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2: BREAKDOWN BY FUND TYPE (formula-driven)
# ═══════════════════════════════════════════════════════════════════════════════
r += 1
write_section_title(ws2, r, "BREAKDOWN BY FUND TYPE", CLR_BREAKDOWN)
r += 1
write_subhdr(ws2, r, ["Fund Type", "Qualified", "100% Rate", "Shariah", "Conventional", "", "", ""], CLR_BREAKDOWN)
r += 1

FUND_TYPES = ["Equity", "Mixed Asset / Balanced", "Fixed Income", "Fund of Funds", "Money Market"]

type_start_row = r
for ft in FUND_TYPES:
    # COUNTIFS formulas referencing Master sheet
    qual_f = f'=COUNTIFS({MASTER_REF}{M_STATUS}${data_start}:{M_STATUS}${last_row},"Qualified",{MASTER_REF}{M_FUND_TYPE}${data_start}:{M_FUND_TYPE}${last_row},"{ft}")'
    perfect_f = f'=COUNTIFS({MASTER_REF}{M_STATUS}${data_start}:{M_STATUS}${last_row},"Qualified",{MASTER_REF}{M_FUND_TYPE}${data_start}:{M_FUND_TYPE}${last_row},"{ft}",{MASTER_REF}{M_BEAT}${data_start}:{M_BEAT}${last_row},100)'
    shariah_f = f'=COUNTIFS({MASTER_REF}{M_STATUS}${data_start}:{M_STATUS}${last_row},"Qualified",{MASTER_REF}{M_FUND_TYPE}${data_start}:{M_FUND_TYPE}${last_row},"{ft}",{MASTER_REF}{M_SERIES}${data_start}:{M_SERIES}${last_row},"Shariah")'
    conv_f = f'=B{r}-D{r}'
    write_formula_row(ws2, r, [ft, qual_f, perfect_f, shariah_f, conv_f],
                      alt=(r % 2 == 0), number_formats=["@", "0", "0", "0", "0"])
    r += 1

# Total row
total_qual_f = f'=SUM(B{type_start_row}:B{r-1})'
total_perf_f = f'=SUM(C{type_start_row}:C{r-1})'
total_shar_f = f'=SUM(D{type_start_row}:D{r-1})'
total_conv_f = f'=SUM(E{type_start_row}:E{r-1})'
write_total_row(ws2, r, ["TOTAL", total_qual_f, total_perf_f, total_shar_f, total_conv_f],
                number_formats=["@", "0", "0", "0", "0"])
r += 2


# ═══════════════════════════════════════════════════════════════════════════════
# PER-FUND-TYPE SECTIONS: Median Stats + Top 5 by 3Y Fund / 3Y Alpha / AE / AUM
# ═══════════════════════════════════════════════════════════════════════════════

def base_cond(fund_type):
    """Returns the filter condition string for Qualified + matching fund type."""
    return f'({MASTER_REF}{M_STATUS}${data_start}:{M_STATUS}${last_row}="Qualified")*({MASTER_REF}{M_FUND_TYPE}${data_start}:{M_FUND_TYPE}${last_row}="{fund_type}")'

def median_f(dcol, fund_type):
    """Returns MEDIAN(FILTER()) formula string."""
    cond = base_cond(fund_type)
    return f'=IFERROR(MEDIAN(FILTER({MASTER_REF}{dcol}${data_start}:{dcol}${last_row},{cond})),"")'

def rank_f(array_cols, sort_col, rank, fund_type, col_idx):
    """
    Google Sheets compatible formula: SORT(FILTER({multi-col array}, cond1, cond2, ...), sort_idx, FALSE).
    - array_cols: ordered list of Master column letters in the FILTER array (no "" entries)
    - sort_col:   column letter to sort by (must be present in array_cols)
    - rank:       1-5 row rank to extract
    - fund_type:  e.g. "Equity"
    - col_idx:    1-based column index within array_cols to extract
    """
    array_parts = [f"{MASTER_REF}{col}${data_start}:{col}${last_row}" for col in array_cols]
    array_str = "{" + ",".join(array_parts) + "}"
    sort_idx = array_cols.index(sort_col) + 1
    status_cond = f'{MASTER_REF}{M_STATUS}${data_start}:{M_STATUS}${last_row}="Qualified"'
    type_cond   = f'{MASTER_REF}{M_FUND_TYPE}${data_start}:{M_FUND_TYPE}${last_row}="{fund_type}"'
    blank_cond  = f'{MASTER_REF}{sort_col}${data_start}:{sort_col}${last_row}<>""'
    return f'=IFERROR(INDEX(SORT(FILTER({array_str},{status_cond},{type_cond},{blank_cond}),{sort_idx},FALSE),{rank},{col_idx}),"")'

def write_top5(ws, r, title, headers, fund_type, sort_col, col_specs, color, nfmts=None):
    """Write a full Top-5 table using Google Sheets compatible SORT(FILTER({array})) formulas.
    col_specs: list of (col_letter, is_sort) tuples — use "" col_letter for filler/empty columns.
    nfmts: per-column number formats (default: standard layout)
    """
    if nfmts is None:
        nfmts = [None, None, "0.00", "0.00", "0.0", "#,##0.00", None, None]
    write_section_title(ws, r, f"{title}", color)
    r += 1
    write_subhdr(ws, r, headers, color)
    r += 1

    # Build the ordered column list for the FILTER array (excludes "" filler entries)
    array_cols = [col for col, _ in col_specs if col != ""]

    for rank_num in range(1, 6):
        formulas = []
        col_idx = 0  # tracks 1-based position within array_cols
        for col_letter, _ in col_specs:
            if col_letter == "":
                formulas.append("")  # empty filler column — no formula
            else:
                col_idx += 1
                f = rank_f(array_cols, sort_col, rank_num, fund_type, col_idx)
                formulas.append(f)

        write_formula_row(ws, r, formulas, alt=((rank_num - 1) % 2 == 0),
                         number_formats=nfmts)
        r += 1

    r += 1
    return r

def build_fund_type_section(ws, r, fund_type, color, fund_type_label=None):
    """Build a complete section for a fund type: Median + 4 Top-5 tables.
    Uses Google Sheets compatible formulas: MEDIAN(FILTER) and SORT(FILTER({array})).
    """
    label = fund_type_label or fund_type

    # ── Median Statistics ──────────────────────────────────────────────────
    write_section_title(ws, r, f"{label} — MEDIAN STATISTICS", color)
    r += 1
    write_subhdr(ws, r, ["Metric", "YTD (%)", "1Y (%)", "3Y Ann. (%)", "5Y Ann. (%)", "10Y Ann. (%)", "3Y Alpha (%)", "Beat (%)"], color)
    r += 1

    # Median row with formulas - map column letters to the data columns
    med_formulas = [
        "Median",
        median_f(M_YTD_FUND, fund_type),   # YTD Fund (col N)
        median_f(M_1Y_FUND, fund_type),    # 1Y Fund (col Q)
        median_f(M_3Y_FUND, fund_type),    # 3Y Fund (col T)
        median_f(M_5Y_FUND, fund_type),    # 5Y Fund (col W)
        median_f(M_10Y_FUND, fund_type),   # 10Y Fund (col Z)
        median_f(M_3Y_ALPHA, fund_type),   # 3Y Alpha (col V)
        median_f(M_BEAT, fund_type),        # Beat % (col L)
    ]
    write_formula_row(ws, r, med_formulas, alt=False, number_formats=["@", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00"])
    r += 2

    # ── Top 5 by 3Y Ann. Fund Return (sort by M_3Y_FUND = col T = 20) ──
    r = write_top5(ws, r, f"{label} — TOP 5 BY 3Y ANN. FUND RETURN",
        ["Fund Name", "Abbr", "3Y Fund (%)", "3Y Alpha (%)", "Beat (%)", "AUM (RM M)", "VC", ""],
        fund_type, M_3Y_FUND,
        [("A", False), ("B", False), ("T", True), ("V", False), ("L", False), ("I", False), (M_VC, False), ("", False)],
        color
    )

    # ── Top 5 by 3Y Ann. Alpha (sort by M_3Y_ALPHA = col V = 22) ──
    r = write_top5(ws, r, f"{label} — TOP 5 BY 3Y ANN. ALPHA",
        ["Fund Name", "Abbr", "3Y Fund (%)", "3Y Alpha (%)", "Beat (%)", "AUM (RM M)", "VC", ""],
        fund_type, M_3Y_ALPHA,
        [("A", False), ("B", False), ("T", False), ("V", True), ("L", False), ("I", False), (M_VC, False), ("", False)],
        color
    )

    # ── Top 5 by Alpha Efficiency (sort by M_AE_3Y = col AE = 31) ──
    r = write_top5(ws, r, f"{label} — TOP 5 BY ALPHA EFFICIENCY",
        ["Fund Name", "Abbr", "Alpha Eff", "3Y Alpha (%)", "Beat (%)", "AUM (RM M)", "VC", ""],
        fund_type, M_AE_3Y,
        [("A", False), ("B", False), ("AE", True), ("V", False), ("L", False), ("I", False), (M_VC, False), ("", False)],
        color
    )

    # ── Top 5 by AUM (sort by M_SIZE = col I = 9) ──
    r = write_top5(ws, r, f"{label} — TOP 5 BY AUM",
        ["Fund Name", "Abbr", "AUM (RM M)", "3Y Fund (%)", "3Y Alpha (%)", "Beat (%)", "VC", ""],
        fund_type, M_SIZE,
        [("A", False), ("B", False), ("I", True), ("T", False), ("V", False), ("L", False), (M_VC, False), ("", False)],
        color,
        nfmts=[None, None, "#,##0.00", "0.00", "0.00", "0.0", None, None]
    )

    return r


# Build sections for each fund type
FUND_TYPE_SECTIONS = [
    ("Equity", CLR_EQUITY, "EQUITY"),
    ("Mixed Asset / Balanced", CLR_MIXED, "BALANCED / MIXED ASSET"),
    ("Fixed Income", CLR_FI, "FIXED INCOME"),
    ("Money Market", CLR_MM, "MONEY MARKET"),
    ("Fund of Funds", CLR_FOF, "FUND OF FUNDS"),
]

for ft, color, label in FUND_TYPE_SECTIONS:
    r = build_fund_type_section(ws2, r, ft, color, label)

# ── Footer ────────────────────────────────────────────────────────────────────
r += 1
ws2.cell(row=r, column=1,
    value="Source: Public Mutual MFR February 2026 | Screened by 5-Point Code Review Framework | Generated April 2026")
ws2.cell(row=r, column=1).font = Font(name="Arial", size=8, italic=True, color="808080")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SUMMARY_COLS)


# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
print(f"Columns: {NUM_COLS}")
print(f"Data rows: {len(rows)} (qualified: {len(qual_rows)}, disqualified: {len(disq_rows_list)})")
